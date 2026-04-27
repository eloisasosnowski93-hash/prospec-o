# ╔══════════════════════════════════════════════════════════════╗
# ║  PROSPEC-O v5  |  Scitec Inteligência Comercial              ║
# ║  Arquivo ÚNICO — produção, deploy-ready, sem módulos extras  ║
# ╚══════════════════════════════════════════════════════════════╝

import re, io, csv, json, os, time, base64, logging, sqlite3, random
from datetime import datetime, timedelta
from unicodedata import normalize

import streamlit as st
import pandas as pd
import requests
from bs4 import BeautifulSoup

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────────────────────────────
def _slug(s):
    if not s: return ""
    s = normalize("NFKD", s).encode("ascii","ignore").decode()
    return re.sub(r"[^a-z0-9]","", s.lower())

RE_EMAIL = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
RE_PHONE = re.compile(r"(?:\(?\d{2}\)?[\s\-]?)(?:9\s?)?\d{4}[\s\-]?\d{4}")
_HEADS   = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Accept-Language":"pt-BR,pt;q=0.9"}

# ─────────────────────────────────────────────────────────────────────
#  DATABASE
# ─────────────────────────────────────────────────────────────────────
class Database:
    def __init__(self, path="prospec_inmetro.db"):
        self.path = path
        self._init()
        try:
            self._migrate()
        except Exception as _e:
            logger.warning(f"migrate: {_e}")

    def _conn(self):
        c = sqlite3.connect(self.path, check_same_thread=False)
        c.row_factory = sqlite3.Row
        return c

    def _init(self):
        with self._conn() as c:
            c.executescript("""
            CREATE TABLE IF NOT EXISTS leads (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                cnpj           TEXT,
                nome           TEXT,
                email          TEXT,
                telefone       TEXT,
                decisor        TEXT,
                ocp            TEXT,
                ocp_cnpj       TEXT,
                validade       TEXT,
                portaria       TEXT,
                estado         TEXT,
                cidade         TEXT,
                situacao       TEXT,
                fonte          TEXT,
                raw_json       TEXT,
                urgencia       TEXT DEFAULT 'normal',
                status_crm     TEXT DEFAULT 'novo',
                data_ult_tent  TEXT,
                recontatar_em  TEXT,
                nao_prospectar INTEGER DEFAULT 0,
                setor          TEXT,
                criado_em      TEXT DEFAULT (datetime('now')),
                atualizado_em  TEXT DEFAULT (datetime('now')),
                UNIQUE(cnpj, portaria)
            );
            CREATE INDEX IF NOT EXISTS idx_p  ON leads(portaria);
            CREATE INDEX IF NOT EXISTS idx_u  ON leads(urgencia);
            CREATE INDEX IF NOT EXISTS idx_sc ON leads(status_crm);
            CREATE TABLE IF NOT EXISTS endotoxina_leads (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                nome        TEXT,
                email       TEXT,
                telefone    TEXT,
                site        TEXT,
                cidade      TEXT,
                estado      TEXT,
                fonte_busca TEXT,
                snippet     TEXT,
                status      TEXT DEFAULT 'novo',
                criado_em   TEXT DEFAULT (datetime('now')),
                UNIQUE(nome, site)
            );
            """)

    def _migrate(self):
        cols = {
            "status_crm":    "TEXT DEFAULT 'novo'",
            "data_ult_tent": "TEXT",
            "recontatar_em": "TEXT",
            "nao_prospectar":"INTEGER DEFAULT 0",
            "setor":         "TEXT",
            "escopo":        "TEXT",       # ISO10993 | MRI | BIOBURDEN | ESTERILIDADE
            "crm_status":    "TEXT",       # novo | vendida | perdida | em_andamento
            "crm_produto":   "TEXT",       # produto no CRM (10993, MRI, etc.)
            "lookalike":     "INTEGER DEFAULT 0",  # score 0-100
        }
        with self._conn() as c:
            existing = {r[1] for r in c.execute("PRAGMA table_info(leads)")}
            for col, td in cols.items():
                if col not in existing:
                    try: c.execute(f"ALTER TABLE leads ADD COLUMN {col} {td}")
                    except: pass
        # Endotoxina: colunas extras
        cols_endo = {"escopo":"TEXT","crm_status":"TEXT","lookalike":"INTEGER DEFAULT 0"}
        with self._conn() as c:
            existing_e = {r[1] for r in c.execute("PRAGMA table_info(endotoxina_leads)")}
            for col, td in cols_endo.items():
                if col not in existing_e:
                    try: c.execute(f"ALTER TABLE endotoxina_leads ADD COLUMN {col} {td}")
                    except: pass

    # ── upsert ───────────────────────────────────────────────────────
    def upsert_lead(self, lead, portaria):
        urg = "normal"
        val = lead.get("validade") or ""
        if val:
            try:
                d = (datetime.strptime(val[:10],"%Y-%m-%d") - datetime.now()).days
                urg = "vencido" if d<0 else "hot" if d<=90 else "medio" if d<=180 else "normal"
            except: pass
        sql = """
        INSERT INTO leads
            (cnpj,nome,email,telefone,decisor,ocp,ocp_cnpj,validade,portaria,
             estado,cidade,situacao,fonte,raw_json,urgencia,setor,
             escopo,crm_status,lookalike,atualizado_em)
        VALUES
            (:cnpj,:nome,:email,:telefone,:decisor,:ocp,:ocp_cnpj,:validade,:portaria,
             :estado,:cidade,:situacao,:fonte,:raw_json,:urgencia,:setor,
             :escopo,:crm_status,:lookalike,datetime('now'))
        ON CONFLICT(cnpj,portaria) DO UPDATE SET
            nome=excluded.nome, email=COALESCE(excluded.email,leads.email),
            telefone=COALESCE(excluded.telefone,leads.telefone),
            decisor=COALESCE(excluded.decisor,leads.decisor),
            ocp=COALESCE(excluded.ocp,leads.ocp), validade=excluded.validade,
            estado=COALESCE(excluded.estado,leads.estado),
            cidade=COALESCE(excluded.cidade,leads.cidade),
            situacao=excluded.situacao, fonte=excluded.fonte,
            setor=COALESCE(excluded.setor,leads.setor),
            escopo=COALESCE(excluded.escopo,leads.escopo),
            crm_status=COALESCE(excluded.crm_status,leads.crm_status),
            lookalike=COALESCE(excluded.lookalike,leads.lookalike),
            raw_json=excluded.raw_json, urgencia=excluded.urgencia,
            atualizado_em=datetime('now')
        WHERE leads.nao_prospectar=0"""
        p = {k: lead.get(k) for k in ("cnpj","nome","email","telefone","decisor","ocp",
                                       "ocp_cnpj","validade","estado","cidade","situacao",
                                       "fonte","setor","escopo","crm_status")}
        p.update({"portaria":portaria,"cnpj":lead.get("cnpj",""),
                  "nome":lead.get("nome") or lead.get("razao_social",""),
                  "raw_json":json.dumps(lead,ensure_ascii=False),"urgencia":urg,
                  "lookalike":lead.get("lookalike",0) or 0})
        with self._conn() as c: c.execute(sql, p)

    # ── queries ──────────────────────────────────────────────────────
    def get_leads(self, f=None):
        where, p = ["nao_prospectar=0"], {}
        if f:
            if f.get("portaria"):  where.append("portaria=:portaria");  p["portaria"]=f["portaria"]
            if f.get("urgencia"):  where.append("urgencia=:urgencia");  p["urgencia"]=f["urgencia"]
            if f.get("status_crm"):where.append("status_crm=:sc");     p["sc"]=f["status_crm"]
            if f.get("estado") not in (None,"","Todos"):
                where.append("estado=:estado"); p["estado"]=f["estado"]
            if f.get("cidade"):
                where.append("LOWER(cidade) LIKE :cidade"); p["cidade"]=f"%{f['cidade'].lower()}%"
            if f.get("ocp"):
                where.append("ocp=:ocp"); p["ocp"]=f["ocp"]
            if f.get("busca"):
                term = f"%{f['busca'].lower()}%"
                where.append("(LOWER(nome) LIKE :busca OR LOWER(cnpj) LIKE :busca OR LOWER(cidade) LIKE :busca)")
                p["busca"]=term
        sql = "SELECT * FROM leads WHERE "+" AND ".join(where)+" ORDER BY urgencia DESC, validade ASC"
        with self._conn() as c:
            return [dict(r) for r in c.execute(sql,p).fetchall()]

    def get_lead_by_id(self, lid):
        with self._conn() as c:
            r = c.execute("SELECT * FROM leads WHERE id=?", (lid,)).fetchone()
        return dict(r) if r else None

    def set_lead_status(self, lid, status):
        now = datetime.now().strftime("%Y-%m-%d")
        rec = (datetime.now()+timedelta(days=180)).strftime("%Y-%m-%d")
        if status=="contatado":
            with self._conn() as c:
                c.execute("UPDATE leads SET status_crm='contatado',data_ult_tent=?,recontatar_em=? WHERE id=?",(now,rec,lid))
        else:
            with self._conn() as c:
                c.execute("UPDATE leads SET nao_prospectar=1,status_crm='nao_prospectar' WHERE id=?",(lid,))

    def get_recontatar(self, portaria=None):
        where = ["status_crm='contatado'","recontatar_em<=:hoje","nao_prospectar=0"]
        p = {"hoje":datetime.now().strftime("%Y-%m-%d")}
        if portaria: where.append("portaria=:portaria"); p["portaria"]=portaria
        with self._conn() as c:
            return [dict(r) for r in c.execute("SELECT * FROM leads WHERE "+" AND ".join(where),p).fetchall()]

    def get_stats(self, portaria=None):
        def _q(extra=None):
            cond,p=[],{}
            if portaria: cond.append("portaria=:portaria"); p["portaria"]=portaria
            if extra: cond.append(extra[0]); p.update(extra[1])
            sql="SELECT COUNT(*) as n FROM leads"+(" WHERE "+" AND ".join(cond) if cond else "")
            with self._conn() as c: return c.execute(sql,p).fetchone()["n"]
        return {"total":_q(),"hot":_q(("urgencia=:u",{"u":"hot"})),
                "vencido":_q(("urgencia=:u",{"u":"vencido"})),
                "medio":_q(("urgencia=:u",{"u":"medio"})),
                "normal":_q(("urgencia=:u",{"u":"normal"})),
                "contatado":_q(("status_crm=:u",{"u":"contatado"}))}

    def get_dist_estado(self, portaria=None):
        cond = "WHERE portaria=? AND nao_prospectar=0" if portaria else "WHERE nao_prospectar=0"
        args = (portaria,) if portaria else ()
        with self._conn() as c:
            rows = c.execute(f"SELECT estado,COUNT(*) as n FROM leads {cond} GROUP BY estado ORDER BY n DESC",args).fetchall()
        return {r["estado"]:r["n"] for r in rows if r["estado"]}

    def get_dist_ocp(self, portaria=None):
        cond = "WHERE portaria=? AND nao_prospectar=0" if portaria else "WHERE nao_prospectar=0"
        args = (portaria,) if portaria else ()
        with self._conn() as c:
            rows = c.execute(f"SELECT ocp,COUNT(*) as n FROM leads {cond} GROUP BY ocp ORDER BY n DESC LIMIT 8",args).fetchall()
        return {r["ocp"]:r["n"] for r in rows if r["ocp"]}

    def clear_leads(self, portaria):
        with self._conn() as c: c.execute("DELETE FROM leads WHERE portaria=?",(portaria,))

    # ── endotoxina ───────────────────────────────────────────────────
    def upsert_endo(self, lead):
        sql = """INSERT OR IGNORE INTO endotoxina_leads
                 (nome,email,telefone,site,cidade,estado,fonte_busca,snippet,escopo,crm_status,lookalike)
                 VALUES (:nome,:email,:telefone,:site,:cidade,:estado,:fonte_busca,:snippet,:escopo,:crm_status,:lookalike)"""
        p = {k:lead.get(k) for k in ("nome","email","telefone","site","cidade","estado","fonte_busca","snippet","escopo","crm_status")}
        p["lookalike"] = lead.get("lookalike",0) or 0
        with self._conn() as c: c.execute(sql,p)

    def set_endo_status(self, lid, status):
        with self._conn() as c:
            c.execute("UPDATE endotoxina_leads SET status=? WHERE id=?",(status,lid))

    def get_endo_leads(self, status_filter=None):
        sql = "SELECT * FROM endotoxina_leads"
        p = {}
        if status_filter: sql+=" WHERE status=:s"; p["s"]=status_filter
        sql+=" ORDER BY criado_em DESC"
        with self._conn() as c: return [dict(r) for r in c.execute(sql,p).fetchall()]

    def get_endo_stats(self):
        with self._conn() as c:
            total = c.execute("SELECT COUNT(*) as n FROM endotoxina_leads").fetchone()["n"]
            novo  = c.execute("SELECT COUNT(*) as n FROM endotoxina_leads WHERE status='novo'").fetchone()["n"]
            cont  = c.execute("SELECT COUNT(*) as n FROM endotoxina_leads WHERE status='contatado'").fetchone()["n"]
        return {"total":total,"novo":novo,"contatado":cont}

    def clear_endo(self):
        with self._conn() as c: c.execute("DELETE FROM endotoxina_leads")


# ─────────────────────────────────────────────────────────────────────
#  CRM READER
# ─────────────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────
#  Palavras-chave lookalike derivadas do perfil de sucesso no CRM
#  (extraídas das empresas com Estado="Vendida" no RD Station)
# ─────────────────────────────────────────────────────────────────────
LOOKALIKE_KWS = [
    # termos de produto/setor das empresas que fecharam negócio
    "implant","ortopéd","ortoped","biomédic","biomédica","odontológ","dental",
    "engenharia médic","equipamentos médic","produto médic","tecnologia médic",
    "biomaterial","biologic","biomed","medical","device",
    # setores CNAE lookalike (indústria de dispositivos médicos)
    "fabricação de equipamentos e instrumentos médicos",
    "fabricação de aparelhos e equipamentos de medição",
    "fabricação de instrumentos odontológicos",
    "comércio de equipamentos médico-hospitalares",
]

# CNAEs mapeados por escopo de ensaio (conforme especificação Scitec)
LOOKALIKE_CNAES = {
    "3250701","3250702","3250703","3250704","3250705","3250706","3250707",
    "2660400",
    "8129000","8650099",
    "4645101","4645102","4773300",
    "2121101","2122600",
}

# Mapeamento preciso: escopo → CNAEs alvo
CNAE_POR_ESCOPO = {
    "ISO10993":  ["3250701"],           # Fab. equip. e instrumentos médico-odontológicos
    "MRI":       ["2660400"],           # Fab. equip. irradiação eletromagnética/imagem médica
    "BIOBURDEN": ["3250701","3250702"], # Fab. equip. médicos + artefatos hospitalar
    "ENDO_EST":  ["8129000","8650099","2121101","2122600"],
    # 8129-0/00 = serviços limpeza/higienização (farmácias de manipulação registram aqui)
    # 8650-0/99 = outros serviços de saúde humana
    # 2121-1/01 = fabricação de medicamentos homeopáticos/manipulados
    # 2122-6/00 = fabricação de medicamentos para uso humano
}

# Padrões que EXCLUEM resultados que NÃO são empresas (artigos, guias, portais, ERPs)
_EXCL_NAO_EMPRESA_RE = [
    r"(?i)como\s+registrar",
    r"(?i)guia\s+(completo|detalhado|prático|passo)",
    r"(?i)passo\s+a\s+passo",
    r"(?i)consulta\s+(a\s+)?registro",
    r"(?i)registro\s+anvisa\s+n[ºo°\.]",
    r"(?i)agência\s+nacional\s+de",
    r"(?i)portal\s+(anvisa|gov|saúde)",
    r"(?i)entenda\s+(o\s+que|como)",
    r"(?i)tudo\s+sobre",
    r"(?i)o\s+que\s+é\s+(a\s+)?iso",
    r"(?i)manual\s+de\s+",
    r"(?i)saiba\s+mais\s+sobre",
    r"(?i)anvisa\.gov\.br",
    r"(?i)smERP|Tecnoprocess|smerp\.com",
    r"(?i)\|\s*(blog|artigo|post|notícia)",
    r"(?i)norma\s+iso\s+\d",
    r"(?i)resolução\s+(rdc|re)\s+\d",
    r"(?i)lei\s+n[ºo]\s+\d",
    r"(?i)publicado\s+em\s+\d",
    r"(?i)atualizado\s+em",
]
import re as _re_mod
_EXCL_NAO_EMPRESA = [_re_mod.compile(p) for p in _EXCL_NAO_EMPRESA_RE]

def _is_empresa(nome: str) -> bool:
    """Retorna True APENAS se o texto é uma razão social de empresa.
    Rejeita: artigos, guias, portais, filtros de busca, nomes de produtos."""
    nome_norm = nome.strip()
    if not nome_norm or len(nome_norm) < 4:
        return False

    # ── Rejeição imediata: padrões específicos ────────────────────────
    nome_up = nome_norm.upper()

    # Padrões Econodata (filtros de busca que aparecem como "empresa")
    _ECONODATA_FILTERS = [
        "EMPRESAS DE BUSCA", "MAIS DE", "FUNCIONÁRIOS EM",
        "FILTRAR POR", "ORDENAR POR", "VER MAIS", "PRÓXIMA PÁGINA",
    ]
    if any(f in nome_up for f in _ECONODATA_FILTERS):
        return False

    # Padrões de conteúdo (artigos, guias, registros)
    for pat in _EXCL_NAO_EMPRESA:
        if pat.search(nome_norm):
            return False

    # Nomes muito longos (>10 palavras = descrição, não empresa)
    palavras = nome_norm.split()
    if len(palavras) > 10:
        return False

    # Começa com verbo ou artigo comum = provavelmente não é empresa
    _NAO_INICIO_EMPRESA = {
        "como","passo","guia","manual","entenda","saiba","tudo","what","how",
        "consulta","registro","norma","resolução","o","a","os","as","um","uma",
        "para","por","sobre","de","em","no","na","nos","nas","ao","aos",
        "portal","acesso","sistema","módulo","serviço","download","veja",
    }
    if palavras and palavras[0].lower() in _NAO_INICIO_EMPRESA:
        return False
    # Rejeita combinações "Portal X" ou "Sistema X" no meio
    if re.search(r"(?i)\b(portal|sistema|plataforma)\s+\w+\s*[-–]", nome_norm):
        return False

    # ── Indicadores positivos de empresa ─────────────────────────────
    _TERMOS_EMPRESA = [
        "LTDA","LTDA.","S.A.","S/A","S.A","EIRELI","ME ","EPP ","ME.","EPP.",
        " IND "," IND."," COM "," COM.","INDÚSTRIA","INDUSTRIA","COMÉRCIO",
        "COMERCIO","IMPORTAÇÃO","EXPORTAÇÃO","TECNOLOGIA","ENGENHARIA",
        "MEDICAL","MÉDICA","BIOMÉDICA","BIOMED","IMPLANTES","ORTOPEDIA",
        "FARMACÊUTICA","FARMÁCIA","DISTRIBUIDORA","FABRICANTE","FABRICAÇÃO",
        "PRODUTOS","SISTEMAS","SOLUÇÕES","EQUIPAMENTOS","INSTRUMENTAL",
    ]
    if any(t in nome_up for t in _TERMOS_EMPRESA):
        return True

    # Heurística final: 2-7 palavras, sem padrão de conteúdo = provável empresa
    return 2 <= len(palavras) <= 7

# Seeds específicas para cada escopo de ensaio
SEEDS_ISO10993 = [
    "fabricante implante ortopédico biocompatibilidade ISO 10993 ANVISA",
    "fabricante dispositivo médico implantável ensaios biológicos ISO 10993",
    "fabricante implante dental odontológico certificação biocompatibilidade",
    "fabricante prótese biomédica material contato tecido ensaio biológico",
    "indústria biomateriais implante cirúrgico avaliação biológica ANVISA",
    "fabricante kit cirúrgico curativo contato prolongado ISO 10993",
    "fabricante válvula cardíaca stent dispositivo implantável biocompatível",
    "empresa biomédica desenvolvimento produto implantável registro ANVISA",
]
SEEDS_MRI = [
    "fabricante implante compatível ressonância magnética MRI ANVISA",
    "fabricante dispositivo médico MRI safe MRI conditional implante",
    "indústria implante ortopédico avaliação artefato MRI ressonância",
    "fabricante produto médico marcação CE MRI safety test implante",
    "empresa dispositivo médico implantável ensaio ressonância magnética",
    "fabricante implante titânio aço inox MRI compatível certificação",
]
SEEDS_BIOBURDEN = [
    "farmácia manipulação injetáveis ANVISA registro Brasil",
    "indústria farmacêutica fabricante produto injetável parenterais Brasil",
    "fabricante dispositivo médico estéril ANVISA registro bioburden",
    "fabricante embalagem primária produto farmacêutico bioburden",
    "fabricante material médico-hospitalar descartável contaminação microbiol",
    "indústria farmacêutica semi-sólido líquido oral parenterais bpf",
]
SEEDS_ESTERILIDADE = [
    "fabricante seringa agulha cateter dreno estéril hospitalar",
    "fabricante implante cateter sutura kit cirúrgico estéril",
    "farmácia compounding manipulação injetável controle microbiológico",
    "fabricante saneante hospitalar alegação esterilidade ANVISA",
    "empresa biotecnologia biofármaco vacina produto biológico Brasil",
    "fabricante colírio solução oftálmica estéril ANVISA",
]


class CRMAnalyzer:
    """Análise e cruzamento com dados do RD Station."""

    @staticmethod
    def load_csv(file_bytes):
        txt = file_bytes.decode("utf-8", errors="ignore")
        if txt.startswith("sep="): txt = "\n".join(txt.splitlines()[1:])
        df = pd.read_csv(io.StringIO(txt), dtype=str).fillna("")
        df.columns = [c.strip() for c in df.columns]
        return df

    @staticmethod
    def empresa_slugs(df):
        col = "Empresa" if "Empresa" in df.columns else df.columns[1]
        return {_slug(v) for v in df[col].tolist() if str(v).strip()}

    @staticmethod
    def summary(df):
        return {"total":len(df),
                "etapas":df["Etapa"].value_counts().to_dict() if "Etapa" in df.columns else {},
                "estados":df["Estado"].value_counts().to_dict() if "Estado" in df.columns else {}}

    # ── Cache interno para performance (evita iterrows repetido) ────
    _crm_cache: dict = {}

    @classmethod
    def _build_cache(cls, df):
        """Constrói cache de busca fuzzy a partir do CRM."""
        if df is None or id(df) in cls._crm_cache:
            return
        index = []
        for _, row in df.iterrows():
            empresa = str(row.get("Empresa","")).strip()
            if not empresa:
                continue
            slug_full = _slug(empresa)
            # Slug normalizado sem sufixos legais (LTDA, SA, EIRELI, etc.)
            slug_short = re.sub(
                r"(ltda|sa|eireli|me|epp|srl|lda|sas|inc|corp|llc|co|cia)"
                r"[\.\s]*$", "", slug_full
            ).rstrip()
            cnpj_raw = re.sub(r"\D","", str(row.get("CNPJ","") or ""))
            index.append({
                "slug_full":  slug_full,
                "slug_short": slug_short,
                "cnpj":       cnpj_raw,
                "estado":     str(row.get("Estado","")).strip(),
                "produto":    str(row.get("Produtos","")).strip(),
                "valor":      str(row.get("Valor Único","")).strip(),
                "contato":    str(row.get("Contatos","")).strip(),
                "cargo":      str(row.get("Cargo","")).strip(),
                "etapa":      str(row.get("Etapa","")).strip(),
                "motivo":     str(row.get("Motivo de Perda","")).strip(),
                "email":      str(row.get("Email","")).strip(),
                "telefone":   str(row.get("Telefone","")).strip(),
                "nome_orig":  empresa,
            })
        cls._crm_cache[id(df)] = index

    @classmethod
    def _find_in_crm(cls, empresa_nome, cnpj_raw, df):
        """Busca empresa no CRM por CNPJ (exato) ou slug (fuzzy).
        Retorna a linha do CRM ou None."""
        if df is None: return None
        cls._build_cache(df)
        index = cls._crm_cache.get(id(df), [])
        if not index: return None

        slug_q = _slug(empresa_nome)
        slug_q_short = re.sub(
            r"(ltda|sa|eireli|me|epp|srl|lda|sas|inc|corp|llc|co|cia)"
            r"[\.\s]*$", "", slug_q
        ).rstrip()
        cnpj_q = re.sub(r"\D","", cnpj_raw or "")

        best = None
        for entry in index:
            # 1. Match exato por CNPJ (máxima confiança)
            if cnpj_q and len(cnpj_q)==14 and entry["cnpj"]==cnpj_q:
                return entry
            # 2. Match exato por slug completo
            if entry["slug_full"] == slug_q:
                best = entry; continue
            # 3. Match por slug sem sufixo legal (fuzzy)
            if (slug_q_short and entry["slug_short"]
                    and len(slug_q_short) >= 6
                    and (slug_q_short.startswith(entry["slug_short"][:8])
                         or entry["slug_short"].startswith(slug_q_short[:8]))):
                if best is None:
                    best = entry
        return best

    @classmethod
    def crm_status(cls, empresa_nome, df, cnpj_raw=""):
        """Retorna status no CRM com fuzzy matching:
        'vendida' | 'perdida' | 'em_andamento' | 'novo'"""
        if df is None: return "novo"
        entry = cls._find_in_crm(empresa_nome, cnpj_raw, df)
        if not entry: return "novo"
        estado = entry["estado"].lower()
        if "vendida"     in estado: return "vendida"
        if "perdida"     in estado: return "perdida"
        if "andamento"   in estado: return "em_andamento"
        return "em_andamento"

    @classmethod
    def crm_detail(cls, empresa_nome, df, cnpj_raw=""):
        """Retorna dict completo com todos os dados do CRM para a empresa."""
        if df is None: return {}
        entry = cls._find_in_crm(empresa_nome, cnpj_raw, df)
        if not entry: return {}
        return {
            "crm_estado":   entry["estado"],
            "crm_produto":  entry["produto"],
            "crm_valor":    entry["valor"],
            "crm_contato":  entry["contato"],
            "crm_cargo":    entry["cargo"],
            "crm_etapa":    entry["etapa"],
            "crm_motivo":   entry["motivo"],
            "crm_email":    entry["email"],
            "crm_telefone": entry["telefone"],
            "crm_nome_orig":entry["nome_orig"],
        }

    @classmethod
    def enrich_lead_from_crm(cls, lead: dict, df) -> dict:
        """Enriquece um lead com dados disponíveis no CRM (contato, cargo, email, telefone)."""
        if df is None: return lead
        entry = cls._find_in_crm(lead.get("nome",""), lead.get("cnpj",""), df)
        if not entry: return lead
        # Preenche campos vazios com dados do CRM
        if entry["email"]    and not lead.get("email"):    lead["email"]    = entry["email"]
        if entry["telefone"] and not lead.get("telefone"): lead["telefone"] = entry["telefone"]
        if entry["contato"]  and not lead.get("decisor"):  lead["decisor"]  = entry["contato"]
        if entry["cargo"]    and not lead.get("cargo_decisor"): lead["cargo_decisor"] = entry["cargo"]
        lead["crm_status"]  = cls.crm_status(lead.get("nome",""), df, lead.get("cnpj",""))
        lead["crm_produto"] = entry["produto"]
        lead["crm_valor"]   = entry["valor"]
        return lead

    @staticmethod
    def lookalike_score(lead, setor_cnae=""):
        """Pontua similaridade com o perfil ICP das empresas vendidas (0-100)."""
        score = 0
        nome_lower  = (lead.get("nome")  or "").lower()
        setor_lower = (lead.get("setor") or setor_cnae or "").lower()
        cnae_lead   = re.sub(r"\D","", lead.get("cnae") or setor_lower)[:7]
        texto = nome_lower + " " + setor_lower

        # Keywords do perfil ICP (LOOKALIKE_KWS definido globalmente)
        for kw in LOOKALIKE_KWS:
            if kw.lower() in texto:
                score += 12

        # CNAE exato ou prefixo de 4 dígitos
        if cnae_lead:
            for c in LOOKALIKE_CNAES:
                if cnae_lead == c or cnae_lead[:4] == c[:4]:
                    score += 30
                    break

        # Bônus: CNAE específico da Scitec
        _CNAES_SCITEC_ALVO = {"3250701","2660400","8129000","8650099"}
        if cnae_lead[:7] in _CNAES_SCITEC_ALVO:
            score += 20

        return min(score, 100)

    @staticmethod
    def success_profile_seeds():
        """Gera seeds lookalike baseadas nas empresas vendidas do CRM."""
        return [
            "fabricante implante ortopédico titânio registro ANVISA Brasil",
            "indústria equipamentos médicos biomédica engenharia Brasil",
            "fabricante implante dental odontológico certificação ANVISA",
            "empresa dispositivo médico implantável biocompatibilidade",
            "indústria biomédica ortopedia equipamentos hospitalares Brasil",
            "fabricante prótese cirúrgica implante biomateriais ANVISA",
        ]

    @classmethod
    def icp_score_lead(cls, lead: dict, df, modulo: str = "LAB") -> dict:
        """Cruzamento completo: lead × Persona ICP do CRM.
        Retorna dict com scores, flags e recomendação de abordagem."""
        nome  = lead.get("nome","")
        cnpj  = lead.get("cnpj","")
        setor = (lead.get("setor") or lead.get("cnae") or "").lower()
        score = {"total": 0, "flags": [], "acao": "prospectar", "prioridade": "normal"}

        # 1. Já existe no CRM?
        crm_status = cls.crm_status(nome, df, cnpj) if df is not None else "novo"
        score["crm_status"] = crm_status
        detail = cls.crm_detail(nome, df, cnpj) if df is not None else {}

        if crm_status == "vendida":
            score["total"] += 30
            score["flags"].append("✅ Cliente ativo — priorizar upsell")
            score["acao"] = "upsell"
            score["prioridade"] = "alta"
        elif crm_status == "perdida":
            motivo = detail.get("crm_motivo","").lower()
            _REATIVACAO_MOTIVOS = ["demanda futura","capacidade de atendimento","sem retorno"]
            if any(m in motivo for m in _REATIVACAO_MOTIVOS):
                score["total"] += 20
                score["flags"].append(f"♻️ Reativar — motivo: {detail.get('crm_motivo','')[:30]}")
                score["acao"] = "reativar"
                score["prioridade"] = "alta"
            else:
                score["flags"].append(f"❌ Perdida — {detail.get('crm_motivo','')[:30]}")
                score["acao"] = "descartar"
        elif crm_status == "em_andamento":
            score["flags"].append("🔄 Em negociação no CRM")
            score["acao"] = "nurturing"

        # 2. Perfil de segmento coincide com ICP?
        _SEG_ICP = ["implant","ortopéd","biomédic","biomaterial","dental","odontológ",
                    "equip médic","produto médic","cirúrg","endoprót","prótese","fixador"]
        seg_matches = [k for k in _SEG_ICP if k in (nome.lower() + setor)]
        if seg_matches:
            score["total"] += len(seg_matches) * 8
            score["flags"].append(f"🏭 Segmento ICP: {seg_matches[0]}")

        # 3. Cargo decisor coincide com cargos de sucesso?
        decisor_cargo = (lead.get("cargo_decisor") or lead.get("decisor") or "").lower()
        _CARGOS_SUCESSO = ["qualidade","regulatório","engenharia","p&d","compras","produção","cq"]
        cargo_matches = [c for c in _CARGOS_SUCESSO if c in decisor_cargo]
        if cargo_matches:
            score["total"] += 15
            score["flags"].append(f"👤 Decisor ICP: {cargo_matches[0]}")

        # 4. CNAE alvo Scitec?
        cnae_lead = re.sub(r"\D","", setor)[:7]
        _CNAES_SCITEC = {"3250701","2660400","8129000","8650099"}
        if cnae_lead[:7] in _CNAES_SCITEC or any(cnae_lead.startswith(c[:4]) for c in _CNAES_SCITEC):
            score["total"] += 20
            score["flags"].append("🎯 CNAE alvo Scitec")

        # 5. Lookalike score base
        look = cls.lookalike_score(lead)
        score["total"] += round(look * 0.3)
        score["lookalike"] = look

        # Prioridade final
        t = min(score["total"], 100)
        score["total"] = t
        if t >= 70 and score["prioridade"] == "normal": score["prioridade"] = "alta"
        elif t >= 40 and score["prioridade"] == "normal": score["prioridade"] = "media"

        return score

# Alias para retrocompatibilidade
CRMReader = CRMAnalyzer


# ─────────────────────────────────────────────────────────────────────
#  SCRAPER INMETRO
# ─────────────────────────────────────────────────────────────────────
class InmetroScraper:
    def __init__(self):
        self._s = requests.Session()
        self._s.headers.update(_HEADS)
        self._s.mount("https://", requests.adapters.HTTPAdapter(max_retries=2))

    def fetch_all_empresas(self, portaria, estado=None, cidade=None):
        for fn in (self._dados_abertos, self._html_inmetro, self._csv_direto):
            try:
                leads = fn(portaria)
                if leads:
                    return self._geo(leads, estado, cidade)
            except Exception as e:
                logger.warning(f"{fn.__name__}: {e}")
        return []

    def _dados_abertos(self, portaria):
        r = self._s.get("https://dados.gov.br/api/3/action/package_search",
                        params={"q":f"inmetro certificacao portaria {portaria}","rows":10},timeout=15)
        r.raise_for_status()
        leads = []
        for ds in r.json().get("result",{}).get("results",[])[:3]:
            for res in ds.get("resources",[]):
                fmt = res.get("format","").upper()
                if fmt in ("CSV","XLSX","JSON"):
                    try: leads.extend(self._dl(res["url"],fmt,portaria))
                    except: pass
        return leads

    def _dl(self, url, fmt, portaria):
        r = self._s.get(url, timeout=30); r.raise_for_status(); leads=[]
        if fmt=="CSV":
            for row in csv.DictReader(io.StringIO(r.content.decode("utf-8","ignore"))):
                l=self._norm(dict(row),portaria)
                if l: leads.append(l)
        elif fmt in ("XLSX","XLS"):
            try:
                import openpyxl
                wb=openpyxl.load_workbook(io.BytesIO(r.content),read_only=True); ws=wb.active
                hdrs=[str(c.value or "").strip().lower() for c in next(ws.iter_rows(max_row=1))]
                for row in ws.iter_rows(min_row=2,values_only=True):
                    l=self._norm(dict(zip(hdrs,row)),portaria)
                    if l: leads.append(l)
            except: pass
        elif fmt=="JSON":
            data=r.json(); items=data if isinstance(data,list) else data.get("results",data.get("data",[]))
            for item in items:
                l=self._norm(item,portaria); 
                if l: leads.append(l)
        return leads

    def _html_inmetro(self, portaria):
        """Scraping com paginação até 20 páginas."""
        base = "https://www.inmetro.gov.br/consumidor/produtos/produtosEspecificacao.asp"
        leads = []
        for page in range(1, 21):
            try:
                params = {"NPortaria": portaria, "pesquisa": "S"}
                if page > 1: params["pagina"] = page
                resp = self._s.get(base, params=params, timeout=20)
                resp.raise_for_status(); resp.encoding = "windows-1252"
                soup = BeautifulSoup(resp.text, "html.parser")
                page_leads = []
                for tbl in soup.find_all("table"):
                    rows = tbl.find_all("tr")
                    if len(rows) < 2: continue
                    hdrs = [th.get_text(strip=True).lower() for th in rows[0].find_all(["th","td"])]
                    if not any(k in " ".join(hdrs) for k in ["empresa","cnpj","razão","fabricante"]): continue
                    for tr in rows[1:]:
                        cells = [td.get_text(strip=True) for td in tr.find_all("td")]
                        if cells:
                            l = self._norm(dict(zip(hdrs,cells)), portaria)
                            if l: page_leads.append(l)
                if not page_leads: break
                leads.extend(page_leads)
                time.sleep(0.4)
            except Exception as e:
                logger.warning(f"Page {page}: {e}"); break
        return leads

    def _csv_direto(self, portaria):
        leads=[]
        for url in [f"https://www.inmetro.gov.br/empresas/ocp/listaocp.csv",
                    f"https://www.inmetro.gov.br/legislacao/upload/portaria_{portaria}_lista.csv"]:
            try:
                r=self._s.get(url,timeout=15)
                if r.status_code==200 and len(r.content)>500:
                    leads.extend(self._dl(url,"CSV",portaria))
            except: pass
        return leads

    def _norm(self, row, portaria):
        def _g(*keys):
            for k in keys:
                for rk,rv in row.items():
                    if k.lower() in str(rk).lower() and rv: return str(rv).strip()
            return None
        cnpj = re.sub(r"\D","",_g("cnpj") or "")
        if len(cnpj)!=14: return None
        c=cnpj
        return {"cnpj":f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}",
                "nome":_g("razão social","razao social","empresa","fabricante","nome"),
                "ocp":_g("ocp","organismo","certificadora"),"ocp_cnpj":_g("cnpj ocp"),
                "validade":_g("validade","vencimento","data validade"),
                "estado":_g("uf","estado"),"cidade":_g("município","municipio","cidade"),
                "situacao":_g("situação","situacao","status"),
                "portaria":portaria,"fonte":"inmetro","email":None,"telefone":None,"decisor":None}

    def _geo(self, leads, estado, cidade):
        if estado and estado!="Todos":
            leads=[l for l in leads if (l.get("estado") or "").upper()==estado.upper()]
        if cidade:
            leads=[l for l in leads if cidade.lower() in (l.get("cidade") or "").lower()]
        return leads

    # ── Demo data expandido: 50 empresas realistas ───────────────────
    def demo_data(self, portaria):
        EMPRESAS_145 = [
            ("Metalúrgica Ferreira Componentes Ltda","11222333000144","SP","São Paulo","Auto Peças"),
            ("Tecno Componentes Automotivos SA","22333444000155","MG","Belo Horizonte","Auto Peças"),
            ("Auto Peças Rápida Importação ME","33444555000166","PR","Curitiba","Importadora"),
            ("Ind. Nacional de Segurança Ltda","44555666000177","RS","Porto Alegre","Equipamentos"),
            ("BrasilParts Comercial e Distrib. Ltda","55666777000188","SC","Joinville","Distribuidora"),
            ("Equipamentos Precisão Industrial Ltda","66777888000199","RJ","Rio de Janeiro","Fabricante"),
            ("Componentes Ouro Preto Fundição SA","77888999000110","MG","Ouro Preto","Fundição"),
            ("Ind. Catarinense de Peças Ltda","88999000000121","SC","Chapecó","Fabricante"),
            ("Grupo Segurança Total Equipamentos Ltda","99000111000132","GO","Goiânia","Distribuidora"),
            ("Fabricação Nordeste Automotiva SA","10111222000143","CE","Fortaleza","Fabricante"),
            ("Peças e Serviços Paulista Ltda","21222333000154","SP","Campinas","Auto Peças"),
            ("Auto Center Sudeste Comércio Ltda","32333444000165","SP","São Bernardo do Campo","Varejo"),
            ("Distribuidora Mineira de Autopeças ME","43444555000176","MG","Uberlândia","Distribuidora"),
            ("Freios e Suspensão do Sul SA","54555666000187","RS","Caxias do Sul","Fabricante"),
            ("Importadora Tech Auto Ltda","65666777000198","SP","São Paulo","Importadora"),
            ("Válvulas e Conexões Brasília Ltda","76777888000109","DF","Brasília","Fabricante"),
            ("Oficina Especializada Motor SA","87888999000120","SP","Santo André","Serviços"),
            ("Peças Originais Recife Ltda","98999000000131","PE","Recife","Distribuidora"),
            ("Auto Elétrica Curitiba ME","19000111000142","PR","Curitiba","Serviços"),
            ("Funilaria e Pintura Premium Ltda","20111222000153","SP","São Paulo","Serviços"),
            ("Borrachas Técnicas do Brasil SA","31222333000164","SP","Sorocaba","Fabricante"),
            ("Vedações e Juntas Sul Ltda","42333444000175","SC","Blumenau","Fabricante"),
            ("Amortecedores BH Comércio Ltda","53444555000186","MG","Contagem","Distribuidora"),
            ("Filtros Automotivos Norte Ltda","64555666000197","PA","Belém","Fabricante"),
            ("Retífica de Motores Leste SA","75666777000108","RJ","Niterói","Serviços"),
            ("Rodas e Pneus Centro-Oeste Ltda","86777888000119","MT","Cuiabá","Varejo"),
            ("Sistema de Freios Avançados SA","97888999000130","SP","Jundiaí","Fabricante"),
            ("Eletrônica Automotiva Gaúcha Ltda","18999000000141","RS","Porto Alegre","Fabricante"),
            ("Acessórios e Tuning Nacional ME","29000111000152","SP","Guarulhos","Varejo"),
            ("Embreagens e Transmissões SP Ltda","30111222000163","SP","Mogi das Cruzes","Fabricante"),
        ]
        EMPRESAS_384 = [
            ("Smith & Allied Medical Devices Ltda","11333444000122","SP","São Paulo","Eletromédico"),
            ("MedTech Brasil Equipamentos SA","22444555000133","SP","São Paulo","Fabricante"),
            ("Philips Medical Brasil Ltda","33555666000144","RJ","Rio de Janeiro","Importadora"),
            ("Siemens Healthineers Brasil Ltda","44666777000155","SP","São Paulo","Importadora"),
            ("Mindray Brasil Distrib. de Equip. Méd.","55777888000166","SP","Campinas","Importadora"),
            ("Dräger Brasil Ltda","66888999000177","SP","São Paulo","Fabricante"),
            ("GE Healthcare Brasil Ltda","77999000000188","SP","São Paulo","Importadora"),
            ("Medtronic do Brasil Ltda","88000111000199","SP","São Paulo","Fabricante"),
            ("Cardiac Science Brasil Ltda","99111222000100","MG","Belo Horizonte","Fabricante"),
            ("Nihon Kohden do Brasil SA","10222333000111","SP","São Paulo","Importadora"),
            ("Biosig Tecnologia Médica Ltda","21333444000122","RS","Porto Alegre","Fabricante"),
            ("Cardiomed Comércio e Distrib. Ltda","32444555000133","PR","Curitiba","Distribuidora"),
            ("Equiphos Equipamentos Hospitalares SA","43555666000144","SP","São Caetano do Sul","Fabricante"),
            ("Instramed Ind. Médico-Hospitalar Ltda","54666777000155","RS","Porto Alegre","Fabricante"),
            ("MedImagem Tecnologia em Saúde Ltda","65777888000166","SP","São Paulo","Fabricante"),
            ("Natus Medical Brasil Ltda","76888999000177","SP","São Paulo","Importadora"),
            ("Orthofix Brasil Ltda","87999000000188","MG","Belo Horizonte","Importadora"),
            ("Philomedical Equipamentos SA","98000111000199","RJ","Rio de Janeiro","Distribuidora"),
            ("QMed Qualidade em Medicina Ltda","19111222000100","CE","Fortaleza","Distribuidora"),
            ("Resmed Brasil Equipamentos Médicos","20222333000111","SP","São Paulo","Importadora"),
        ]
        empresas = EMPRESAS_145 if portaria=="145" else EMPRESAS_384
        ocps = ["Bureau Veritas","SGS Brasil","TÜV Rheinland","DNV GL","Imetro Cert",
                "ABNT","Falcão Bauer","LACTEC","IPT","SENAI"]
        decisores = ["Ana Paula Rodrigues","Carlos Eduardo Lima","Fernanda Souza",
                     "Roberto Mendes","Patricia Alves","João Carlos Neves",
                     "Marcela Oliveira","Diego Santos","Luciana Costa","Paulo Henrique"]
        setores = ["Qualidade","Compras","Engenharia","Diretoria","Certificação"]
        dominios = ["gmail.com","yahoo.com.br","hotmail.com","empresa.com.br","corp.com.br"]

        hoje = datetime.now(); leads=[]
        for nome,cnpj,uf,cidade,setor_emp in empresas:
            c=cnpj; cnpj_fmt=f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}"
            dias = random.randint(-60, 400)
            val = (hoje+timedelta(days=dias)).strftime("%Y-%m-%d")
            decisor = random.choice(decisores)
            slug_emp = re.sub(r"[^a-z]","",nome.lower().split()[0])
            email_gen = f"{slug_emp}@{random.choice(dominios)}"
            fone_ddd = random.choice(["11","21","31","41","51","61","71","85"])
            fone_num = f"9{random.randint(1000,9999)}-{random.randint(1000,9999)}"
            leads.append({
                "cnpj":cnpj_fmt,"nome":nome,"email":email_gen,
                "telefone":f"({fone_ddd}) {fone_num}",
                "decisor":decisor,"ocp":random.choice(ocps),"validade":val,
                "estado":uf,"cidade":cidade,"situacao":"Vigente",
                "portaria":portaria,"fonte":"demo","setor":setor_emp,
            })
        random.shuffle(leads)
        return leads


# ─────────────────────────────────────────────────────────────────────
#  ENRICHER
# ─────────────────────────────────────────────────────────────────────
class DataEnricher:
    def __init__(self):
        self._s = requests.Session()
        self._s.headers.update({"User-Agent":"Mozilla/5.0","Accept":"application/json"})

    # Cargos de decisor mapeados do perfil CRM da Scitec
    _DECISOR_CARGOS = re.compile(
        r"(?i)(gerente|diretor|diretora|coordenador|coordenadora|superintendente|"
        r"vice.president|vp\s|head\s|manager|responsável|chefe\s"
        r"|engenheiro chefe|eng\.?\s*chefe|cto|ceo|coo|cfo|cmo"
        r"|gerência|diretoria|assuntos regulatórios|regulatório|qualidade"
        r"|p&d|pesquisa e desenvolvimento|manufatura|produção|supply)",
        re.UNICODE
    )
    _RE_FONE_BR = re.compile(
        r"(?:(?:(?:\+?55)[\s\-]?)?\(?([1-9]{2})\)?[\s\-]?)"
        r"(?:(?:9[1-9]\d{3}|[2-8]\d{3})[\s\-]?\d{4})"
    )
    _RE_EMAIL_CORP = re.compile(
        r"[a-zA-Z0-9][a-zA-Z0-9._%+\-]{1,40}@"
        r"(?!gmail|yahoo|hotmail|outlook|bol|uol|terra|ig\.)"
        r"[a-zA-Z0-9.\-]{2,40}\.[a-zA-Z]{2,6}"
    )

    def enrich_receita(self, lead):
        cnpj = re.sub(r"\D","",lead.get("cnpj") or "")
        if len(cnpj)!=14 or lead.get("fonte")=="demo": return lead
        # BrasilAPI
        try:
            r=self._s.get(f"https://brasilapi.com.br/api/cnpj/v1/{cnpj}",timeout=10)
            if r.status_code==200:
                d=r.json()
                lead["nome"]     = lead.get("nome") or d.get("razao_social","")
                lead["cidade"]   = lead.get("cidade") or d.get("municipio","")
                lead["estado"]   = lead.get("estado") or d.get("uf","")
                lead["situacao"] = d.get("descricao_situacao_cadastral",lead.get("situacao"))
                # Telefone — prefer telefone2 se for celular (9 dígitos)
                for campo in ("ddd_telefone_1","ddd_telefone_2"):
                    tel = re.sub(r"\D","",d.get(campo) or "")
                    if len(tel)>=10 and not lead.get("telefone"):
                        lead["telefone"]=f"({tel[:2]}) {tel[2:7]}-{tel[7:]}" if len(tel)==11 else f"({tel[:2]}) {tel[2:]}"
                # Email corporativo preferido
                email_raw = (d.get("email") or "").lower().strip()
                if email_raw and not lead.get("email"):
                    # prefer corporate, skip if ends with webmail
                    if not any(wm in email_raw for wm in ("gmail","yahoo","hotmail","outlook","bol","uol")):
                        lead["email"] = email_raw
                    else:
                        lead["email"] = email_raw  # store anyway as fallback
                # QSA — prioriza cargos de decisor
                qsa = d.get("qsa") or []
                best_decisor = None
                for s in qsa:
                    qual = (s.get("descricao_qualificacao_socio") or s.get("qual") or "").lower()
                    nome = s.get("nome_socio") or s.get("nome") or ""
                    cod  = str(s.get("codigo_qualificacao_socio",""))
                    # Sócio-administrador, Diretor, Gerente (códigos RF)
                    if cod in ("5","10","49","22","16","21"):
                        best_decisor = nome; break
                    if self._DECISOR_CARGOS.search(qual) and not best_decisor:
                        best_decisor = nome
                if not best_decisor and qsa:
                    best_decisor = qsa[0].get("nome_socio") or qsa[0].get("nome","")
                if best_decisor and not lead.get("decisor"):
                    lead["decisor"] = best_decisor
                # CNAE e lookalike score
                atv = (d.get("cnae_fiscal_descricao") or "")
                if atv: lead["setor"] = atv[:60]
                lead["lookalike"] = CRMAnalyzer.lookalike_score(lead, atv)
                time.sleep(0.2); return lead
        except: pass
        # receitaws fallback
        try:
            r2=self._s.get(f"https://receitaws.com.br/v1/cnpj/{cnpj}",timeout=10)
            if r2.status_code==200:
                d2=r2.json()
                if d2.get("status")!="ERROR":
                    lead["nome"]   = lead.get("nome") or d2.get("nome","")
                    lead["cidade"] = lead.get("cidade") or d2.get("municipio","")
                    lead["estado"] = lead.get("estado") or d2.get("uf","")
                    tel_raw = re.sub(r"\D","",d2.get("telefone") or "")
                    if len(tel_raw)>=10 and not lead.get("telefone"):
                        lead["telefone"]=f"({tel_raw[:2]}) {tel_raw[2:7]}-{tel_raw[7:]}" if len(tel_raw)==11 else f"({tel_raw[:2]}) {tel_raw[2:]}"
                    if d2.get("email") and not lead.get("email"):
                        lead["email"]=d2["email"].lower()
                    for s in (d2.get("qsa") or []):
                        qual=(s.get("qual") or "").lower()
                        if "admin" in qual or "dir" in qual or self._DECISOR_CARGOS.search(qual):
                            lead["decisor"]=s.get("nome",""); break
                    if not lead.get("decisor") and d2.get("qsa"):
                        lead["decisor"]=d2["qsa"][0].get("nome","")
                    atv2 = d2.get("atividade_principal",[])
                    if isinstance(atv2,list) and atv2:
                        lead.setdefault("setor", atv2[0].get("text","")[:60])
                    lead.setdefault("lookalike", CRMAnalyzer.lookalike_score(lead, lead.get("setor","")))
                time.sleep(0.35)
        except: pass
        return lead

    def enrich_contatos(self, lead):
        if lead.get("email") and lead.get("telefone") and lead.get("decisor"): return lead
        if lead.get("fonte")=="demo": return lead
        cnpj = re.sub(r"\D","",lead.get("cnpj") or "")
        if not lead.get("email"):
            dom = self._dom(lead.get("nome") or "")
            if dom: lead["email"]=f"contato@{dom}"
        try:
            r=self._s.get(f"https://www.cnpj.biz/{cnpj}",timeout=8,
                          headers={"User-Agent":"Mozilla/5.0","Accept":"text/html"})
            if r.status_code==200:
                text = r.text
                # Email corporativo primeiro
                corp_emails=[e.lower() for e in self._RE_EMAIL_CORP.findall(text)
                             if not e.lower().endswith((".png",".jpg",".gif","@sentry.io"))
                             and len(e)<80]
                all_emails=[e.lower() for e in RE_EMAIL.findall(text)
                            if not e.endswith((".png",".jpg",".gif","@sentry.io")) and len(e)<80]
                best_email = corp_emails[0] if corp_emails else (all_emails[0] if all_emails else None)
                if best_email and not lead.get("email"): lead["email"]=best_email
                # Telefone com regex melhorado
                phones = self._RE_FONE_BR.findall(r.text) or RE_PHONE.findall(r.text)
                if phones and not lead.get("telefone"): lead["telefone"]=phones[0]
                # Decisor por cargo na página
                if not lead.get("decisor"):
                    soup_text = BeautifulSoup(text,"html.parser").get_text(" ", strip=True)
                    # Padrão: "Nome Sobrenome - Cargo" ou "Cargo: Nome"
                    pat = re.compile(
                        r"([A-ZÁÉÍÓÚÂÊÎÔÛÃÕÀÈÌÒÙÇ][a-záéíóúâêîôûãõàèìòùç]+(?:\s[A-ZÁÉÍÓÚÂÊÎÔÛÃÕÀÈÌÒÙÇ][a-záéíóúâêîôûãõàèìòùç]+){1,3})"
                        r"\s*[-–|]\s*"
                        r"(gerente|diretor|coordenador|engenheiro|analista|responsável|qualidade|regulatório)",
                        re.UNICODE | re.IGNORECASE
                    )
                    m = pat.search(soup_text)
                    if m: lead["decisor"] = m.group(1).strip()
            time.sleep(0.2)
        except: pass
        return lead

    def _dom(self, nome):
        stop={"ltda","sa","me","eireli","epp","industria","comercio","nacional","brasil",
              "grupo","do","da","de","dos","das","e","imp","exp","distrib","comercial"}
        pts=[p for p in re.sub(r"[^a-zA-Z0-9\s]","",nome.lower()).split() if p not in stop and len(p)>2]
        if not pts: return None
        return (pts[0] if len(pts)==1 else pts[0]+pts[1])+".com.br"


# ─────────────────────────────────────────────────────────────────────
#  ENDOTOXINA BOT
# ─────────────────────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════
#  SEEDS — Foco em CLIENTES que precisam de ensaios
#  A Scitec É o laboratório. Busca quem CONTRATA endotoxina,
#  esterilidade e bioburden — fabricantes e produtores, não concorrentes.
# ══════════════════════════════════════════════════════════════════════

# Padrão (portarias INMETRO) — mostra opções ao usuário
ENDO_SEEDS = [
    "farmácia manipulação injetáveis ANVISA registro Brasil",
    "indústria farmacêutica fabricante produto injetável parenterais Brasil",
    "fabricante dispositivo médico estéril ANVISA registro Brasil",
    "fabricante seringa agulha cateter estéril hospitalar descartável",
    "fabricante embalagem primária produto farmacêutico bioburden",
    "empresa biotecnologia biofármaco vacina produto biológico Brasil",
    "fabricante implante ortopédico prótese cirúrgica estéril ANVISA",
    "fabricante colírio solução oftálmica estéril ANVISA",
    "farmácia compounding manipulação injetável controle microbiológico",
    "indústria farmacêutica fabricante ampola frasco parenterais",
    "fabricante material sutura kit cirúrgico estéril hospitalar",
    "fabricante produto cosmético estéril alegação esterilidade ANVISA",
]

# Modo ENDO_EST (sidebar) — mesmas sementes, porém pre-selecionadas
ENDO_EST_SEEDS = [
    "farmácia manipulação injetáveis ANVISA registro Brasil",
    "indústria farmacêutica fabricante produto injetável parenterais Brasil",
    "fabricante dispositivo médico estéril implantável ANVISA registro",
    "fabricante seringa agulha cateter dreno estéril hospitalar",
    "fabricante embalagem blister frasco produto farmacêutico bioburden",
    "empresa biotecnologia biofármaco vacina produto biológico Brasil",
    "fabricante implante ortopédico prótese cirúrgica estéril ANVISA",
    "fabricante colírio solução oftálmica estéril ANVISA",
    "farmácia compounding manipulação injetável controle microbiológico",
    "fabricante material sutura kit cirúrgico estéril descartável",
    "fabricante saneante hospitalar alegação esterilidade ANVISA",
    "indústria farmacêutica semi-sólido líquido oral parenterais bpf",
]

# Palavras que indicam CLIENTE (fabricante / produtor que precisa do serviço)
ENDO_KWS = [
    "fabricant","indústria","farmácia manipul","compounding","produt","manufat",
    "injetáv","parenteral","ampola","frasco","implant","cateter","sutura",
    "seringa","agulha","estéril","esterilic","colírio","oftálmic","biológ",
    "vacina","biotecnolog","biofárm","cosmét","saneant",
    "anvisa","registro","rdc","bpf","validação","qualidade",
    "bioburden","biocarga","embalag","blister","hospitalar","cirúrg","descartáv",
]

# Palavras que EXCLUEM concorrentes (outros laboratórios prestadores)
ENDO_EXCL = [
    "nossos serviços analíticos","acreditação abnt","prestação serviços laboratorial",
    "nossa metodologia lal","terceiriz","análises terceiros",
]

class EndotoxinaBot:
    """CNAE-FIRST: Casa dos Dados → Econodata → DDG (empresa-only).
    Não retorna artigos, guias, portais ou sites de conteúdo."""

    def __init__(self):
        self._s = requests.Session()
        self._s.headers.update(_HEADS)

    # ── 1. PRIMARY: Casa dos Dados API por CNAE ───────────────────────
    def buscar_cnae(self, cnae: str, uf: str = "SP", paginas: int = 3) -> list:
        """Busca empresas reais pelo CNPJ/CNAE — retorna razão social, telefone, email."""
        resultados = []
        cnae_int = int(re.sub(r"\D","", cnae)) if re.sub(r"\D","", cnae) else 0
        for page in range(1, paginas + 1):
            try:
                r = self._s.post(
                    "https://api.casadosdados.com.br/v2/public/cnpj/search",
                    json={"query": {"cnae_fiscal": cnae_int, "uf": uf.upper()}, "page": page},
                    timeout=15
                )
                if r.status_code != 200: break
                items = r.json().get("data", [])
                if not items: break
                for item in items:
                    c = re.sub(r"\D","", item.get("cnpj",""))
                    cnpj_fmt = f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}" if len(c)==14 else c
                    nome = (item.get("razao_social","") or item.get("nome_fantasia","")).strip()
                    if not nome or not _is_empresa(nome): continue
                    d1 = re.sub(r"\D","", item.get("ddd1","") + item.get("telefone1",""))
                    tel = (f"({d1[:2]}) {d1[2:7]}-{d1[7:]}" if len(d1)==11 else
                           f"({d1[:2]}) {d1[2:]}" if len(d1)==10 else None)
                    resultados.append({
                        "cnpj": cnpj_fmt, "nome": nome,
                        "email": (item.get("email","") or "").lower() or None,
                        "telefone": tel, "site": None,
                        "cidade": item.get("municipio",""), "estado": item.get("uf",""),
                        "setor": item.get("cnae_fiscal_descricao","")[:60],
                        "cnae": cnae, "snippet": item.get("cnae_fiscal_descricao",""),
                        "fonte": "casa_dados", "decisor": None,
                    })
                time.sleep(0.3)
            except Exception as e:
                logger.warning(f"CasaDados CNAE {cnae}/{uf} p{page}: {e}"); break
        return resultados

    # ── 2. SECONDARY: Econodata por CNAE ─────────────────────────────
    def buscar_econodata(self, cnae: str, uf: str = "sp") -> list:
        """Scraping Econodata por CNAE + UF."""
        cnae_c = re.sub(r"\D","", cnae)
        try:
            r = self._s.get(f"https://www.econodata.com.br/empresas/{uf.lower()}/{cnae_c}/1",
                             timeout=15)
            if r.status_code != 200: return []
            soup = BeautifulSoup(r.text, "html.parser")
            cards = (soup.select(".company-card") or soup.select(".empresa-item") or
                     soup.select("article.empresa") or soup.select("tr.empresa"))
            results = []
            for card in cards[:25]:
                nome_el = (card.select_one(".company-name") or card.select_one("h2") or
                           card.select_one("h3") or card.select_one("td"))
                if not nome_el: continue
                nome = nome_el.get_text(strip=True)
                if not nome or not _is_empresa(nome): continue
                cnpj_el = card.select_one(".cnpj")
                cnpj_raw = re.sub(r"\D","", cnpj_el.get_text() if cnpj_el else "")
                results.append({
                    "cnpj": cnpj_raw, "nome": nome, "email": None, "telefone": None,
                    "cidade": "", "estado": uf.upper(), "cnae": cnae_c,
                    "site": None, "snippet": f"CNAE {cnae_c} — {uf.upper()}",
                    "fonte": "econodata", "decisor": None,
                })
            return results
        except Exception as e:
            logger.warning(f"Econodata {cnae}/{uf}: {e}"); return []

    # ── 3. FALLBACK: DDG só para empresas (com filtro rígido) ─────────
    def _ddg(self, query: str, n: int) -> list:
        """DDG HTML — filtra APENAS empresas, bloqueia artigos/guias/portais."""
        query_emp = query + ' "LTDA" OR "S.A." OR "indústria" OR "fabricante" OR "biomédica"'
        try:
            r = self._s.post("https://html.duckduckgo.com/html/",
                              data={"q": query_emp, "kl": "br-pt"}, timeout=15)
            r.raise_for_status()
        except Exception as e:
            logger.warning(f"DDG: {e}"); return []

        soup = BeautifulSoup(r.text, "html.parser")
        results = []
        for res in soup.select(".result")[:n * 3]:
            title_el = res.select_one(".result__title")
            url_el   = res.select_one(".result__url")
            snip_el  = res.select_one(".result__snippet")
            if not title_el: continue
            title   = title_el.get_text(strip=True)
            site    = url_el.get_text(strip=True)  if url_el  else ""
            snippet = snip_el.get_text(strip=True) if snip_el else ""
            # Rejeita: não é empresa
            if not _is_empresa(title): continue
            # Rejeita: concorrente
            if any(ex in (title+snippet).lower() for ex in ENDO_EXCL): continue
            emails = RE_EMAIL.findall(snippet)
            phones = RE_PHONE.findall(snippet)
            results.append({
                "cnpj": None, "nome": title[:120], "site": site[:200],
                "snippet": snippet[:300],
                "email":    emails[0].lower() if emails else None,
                "telefone": phones[0] if phones else None,
                "cidade": None, "estado": None, "fonte": "ddg",
            })
            if len(results) >= n: break
        return results

    def enriquecer_site(self, lead):
        site = lead.get("site","")
        if not site or (lead.get("email") and lead.get("telefone")): return lead
        try:
            url = site if site.startswith("http") else f"https://{site}"
            r   = self._s.get(url, timeout=10)
            if r.status_code == 200:
                emails = [e for e in RE_EMAIL.findall(r.text)
                          if not e.endswith((".png",".jpg",".gif","@sentry"))]
                phones = RE_PHONE.findall(r.text)
                if emails and not lead.get("email"):    lead["email"]    = emails[0].lower()
                if phones and not lead.get("telefone"): lead["telefone"] = phones[0]
        except: pass
        return lead



# ─────────────────────────────────────────────────────────────────────
#  EXPORTER
# ─────────────────────────────────────────────────────────────────────
class DataExporter:
    def __init__(self, db):
        self.db = db

    def _wb_style(self, ws, cols, hcolor="0D9291"):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            hf=Font(bold=True,color="FFFFFF",name="Calibri",size=11)
            hfill=PatternFill("solid",fgColor=hcolor)
            center=Alignment(horizontal="center",vertical="center")
            left=Alignment(horizontal="left",vertical="center",wrap_text=True)
            thin=Side(style="thin",color="DDDDDD"); brd=Border(left=thin,right=thin,top=thin,bottom=thin)
            rfill=PatternFill("solid",fgColor="FDEAEA"); afill=PatternFill("solid",fgColor="FFF8E1")
            ws.row_dimensions[1].height=26
            for ci,(t,w) in enumerate(cols,1):
                c=ws.cell(row=1,column=ci,value=t)
                c.font=hf; c.fill=hfill; c.alignment=center; c.border=brd
                ws.column_dimensions[get_column_letter(ci)].width=w
            ws.freeze_panes="A2"
            return left, brd, rfill, afill
        except: return None,None,None,None

    def generate_excel(self, portaria):
        leads=self.db.get_leads({"portaria":portaria})
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb=openpyxl.Workbook(); ws=wb.active; ws.title=f"P{portaria}"
            cols=[("Nome",38),("CNPJ",20),("E-mail",30),("Telefone",18),("Decisor",28),
                  ("OCP",22),("Validade",14),("Estado",8),("Cidade",20),
                  ("Urgência",12),("Status CRM",14),("Setor",20),("Portaria",10)]
            left,brd,rfill,afill = self._wb_style(ws, cols)
            fields=["nome","cnpj","email","telefone","decisor","ocp","validade",
                    "estado","cidade","urgencia","status_crm","setor","portaria"]
            for ri,lead in enumerate(leads,2):
                for ci,f in enumerate(fields,1):
                    c=ws.cell(row=ri,column=ci,value=lead.get(f,"") or "")
                    if brd: c.border=brd
                    if left: c.alignment=left
                    urg=lead.get("urgencia","")
                    if rfill and urg=="vencido": c.fill=rfill
                    elif afill and urg=="medio": c.fill=afill
            # resumo
            ws2=wb.create_sheet("Resumo"); st2=self.db.get_stats(portaria)
            for ri,(k,v) in enumerate([("Portaria",portaria),("Total",st2["total"]),
                ("Quentes 🔥",st2["hot"]),("Vencidos",st2["vencido"]),
                ("Contatados",st2["contatado"]),("Gerado em",datetime.now().strftime("%d/%m/%Y %H:%M"))],1):
                ws2.cell(row=ri,column=1,value=k).font=Font(bold=True)
                ws2.cell(row=ri,column=2,value=v)
            buf=io.BytesIO(); wb.save(buf); return buf.getvalue()
        except ImportError:
            buf=io.StringIO()
            w=csv.DictWriter(buf,extrasaction="ignore",
                fieldnames=["nome","cnpj","email","telefone","decisor","ocp","validade","estado","cidade","urgencia","status_crm"])
            w.writeheader(); w.writerows(leads)
            return buf.getvalue().encode("utf-8-sig")

    def generate_endo_excel(self):
        leads=self.db.get_endo_leads()
        try:
            import openpyxl
            from openpyxl.styles import Font
            wb=openpyxl.Workbook(); ws=wb.active; ws.title="Endotoxina"
            cols=[("Nome/Empresa",40),("E-mail",30),("Telefone",18),("Site",35),
                  ("Estado",8),("Cidade",20),("Escopo",12),("CRM Status",14),("Lookalike",10),
                  ("Termo Busca",30),("Snippet",45),("Status",12)]
            left,brd,_,_ = self._wb_style(ws, cols, hcolor="7C3AED")
            fields=["nome","email","telefone","site","estado","cidade","escopo","crm_status","lookalike","fonte_busca","snippet","status"]
            for ri,lead in enumerate(leads,2):
                for ci,f in enumerate(fields,1):
                    c=ws.cell(row=ri,column=ci,value=lead.get(f,"") or "")
                    if brd: c.border=brd
                    if left: c.alignment=left
            buf=io.BytesIO(); wb.save(buf); return buf.getvalue()
        except ImportError:
            buf=io.StringIO()
            w=csv.DictWriter(buf,extrasaction="ignore",
                fieldnames=["nome","email","telefone","site","estado","cidade","fonte_busca","snippet","status"])
            w.writeheader(); w.writerows(leads)
            return buf.getvalue().encode("utf-8-sig")


# ─────────────────────────────────────────────────────────────────────
#  LOGO
# ─────────────────────────────────────────────────────────────────────
def _logo_b64():
    # Embedded Scitec logo (inline — funciona em qualquer ambiente)
    return "data:image/png;base64,/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCACFALoDASIAAhEBAxEB/8QAHQABAQACAwEBAQAAAAAAAAAAAAcGCQQFCAMCAf/EAE8QAAAEBAIDDAcDCAYLAAAAAAABAgMEBQYRByESGDEIEyJBUVVWYZSl09QUMjhxgYS0FZGhFjZCcnOxsrMXN3R1ksEjNUZSVGLD0dLh8f/EABoBAQACAwEAAAAAAAAAAAAAAAAEBQECAwb/xAApEQEAAgIBAwMEAQUAAAAAAAAAAQIDEQQSITEFQYETUWHwFCJxkbHh/9oADAMBAAIRAxEAPwDudy5gJhPWmBNO1NU1KenzaM9K9IiPtCKb09CKdQngocJJWSlJZEWzlFM1XMCeg3e0b4wbiX2YqR+d+tfFmARnVcwJ6Dd7RvjBquYE9Bu9o3xhZgARnVcwJ6Dd7RvjBquYE9Bu9o3xhZgARnVcwJ6Dd7RvjBquYE9Bu9o3xhZgARnVcwJ6Dd7RvjBquYE9Bu9o3xhZgARnVcwJ6Dd7RvjBquYE9Bu9o3xhZgARnVcwJ6Dd7RvjBquYE9Bu9o3xhZgARnVcwJ6Dd7RvjBquYE9Bu9o3xhZgARnVcwJ6Dd7RvjBquYE9Bu9o3xhZgARnVcwJ6Dd7RvjBquYE9Bu9o3xhZgARnVcwJ6Dd7RvjBquYE9Bu9o3xhZgARnVcwJ6Dd7RvjDWaNzA0zgNmW4l9mKkfnfrXxZhGdxL7MVI/O/WvizAAAAAAAAAAAAAJjj/X0zomVS9iTttFGTE3SS+4nSJlKNG5kWw1cMrXuWR5DalZvOoceRnpx8c5L+IZnV1VyClID0ueTFqGSZHvbfrOOHyJSWZ/uLjsJKxuioFU5Jt6nH25Yarb6T5KeIuJWha3LlpfEYjRGF1V4gxKKiqSYPQ8DE8P0l9W+Pvp4tBJ7E8hnYtliMhVXMLcMH4JVLNIhkTFtBLNaIojjU/85le9s9hp0erYJHTip2t3lUfX5/K/rxRFK+2/Ms4pmpZHUkt+0ZLMWItgvXNJ2U2e2yknmk/eJ/iDjdTtPrVByRKZ3HJOyjbcsw37156R9Sbl1kJNX+EtVUgURFS1b0zlSkmlbsMRkskcjiCzt1lcss7DqsIncP25upNcw0SslmRML0j9HTf/AHyTwvjcy5SLaN64Ka6oncOGb1TldUYLVjHafefHx+yuOHmNdO1EpEFOCTJZgo7JJ1d2HOSy8rH1Kt1GYqZGRkRkZGR7DIRGtcEZBPYJM1oeMZhFrRpIa3w3IZ4rfoqzNJny5l1EMHo+tK4w2qRimJyhx2EJ1CFwcSrT0EKOxKaWR5FbYRGaeoc5xUv3xz8JNOdn4sxTl13E+LR4+f34epgABGXYAAAAAAA0zjcwNM4DZluJfZipH53618WYRncS+zFSPzv1r4swAAAAAAAAAAAIDuv/APZf5v8A6Ivwh260lUwi5ZIplDQjr0LBKiExLiEmZNae96JqtsLgHns2co7cedZIVnrFZtwrxH4/3DvJnN46R7meGmUudNmKRKIVtDhZGjfDbQZkfEZEo7HxHYTTAXDxFYPv1LMptFstwkVoaMOs0vOOWJRqNy90lwuLM7nmQ+UZidL5pglEUVFwbsNMGIeHZh3E8Jt5LbrZ58aT0UmfJlt4hQNyZ+Y80/vI/wCU2O8xbHjtPidqul8XM5eKu+qsV8flk9X4s0bS8wblkRGOxsQSybeTCFvu8FxmtRntLjIjNXUOjqXDmhcSZaqe0zGQ8JFu5+kwhXbWraZOt5WVnnsVnnceesQJZMJXWM0YmMG9CuLinXEE4gy0kGszJRcpHykPQG5Zlcxl9HzB6Ogn4ZEVFE4wbqDTviNAi0iI+LrGLY4xUi9Z7t+PzLc7kW4+ekTXv8aYRuaY2ay3EeNptcWs4QmXieYJRm3viFEWkkj2Htz4yMcbdDf10Qn7GG/iMVyNVh5hWiKmbxNMzCMUt1R332Lf0lXNKS4k3tyJ5eUQWpJ5G4lYpwsdK5Q8lS1tNNMIPTUSEK9dRkVi23PiLl4xtjnrvN9dtOPKpHH41eLNt26t689nr0AAQXqgAAAAAABpnG5gaZwGzLcS+zFSPzv1r4swjO4l9mKkfnfrXxZgAAAAAAAAAAAfxREpJpURGRlYyPjH9ABD8dsLacYpuYVVJ2fs2KhUk44w0VmXSNREfB/RPPisXVxjl7kz8x5p/eR/ymxmWOBGrCioCSRmfo5Hl+ukSncz1vT0jl0ZIJxGpgYiJi9+ZdeyaVdCU6Jq2JPg8diz2iVE2vhn3UOSmHj+pUmNV3E/5UOGxGw7n7kQzPVQMNFSt9Z71MmknoqQZlptGZGRnlkRcLqGA4hY8RMS4uW0TDqbSo9D055u7i+L/RoPZ71XPqIdrWeAkNNp07M5HPPRGopanXWn299JKlHczQojLLPYf3jvJVTeHWEUuTM5nEtuTA08GJiCJb7h8jTZer7y+JhH0o7x3n7M5P5991vMUr72+8fv9kIw+kL9e4itSyezGMS4+bjkU8s9J49BJmZXVsPK2ezkHrCkqWkNKy8oKRy9qFQfrrtdxw+VSjzP/LiHnjc+xKI/GyIjmmVG2+3FPJ0iK7ZKO5Gf32y5R6hDlWnq6fY9Dw4/pTk1u257+4AAIq9AAAAAAAGmcbmBpnAbMtxL7MVI/O/WvizCM7iX2YqR+d+tfFmAAAAAAAAAAAAAAH5dQh1tTbiErQorKSorkZchkI9iLgXKJsTkdSy0SqNPP0dV/R1n1WzR8Ll1CxgN6XtSdxKPyOLi5FenJG3kpiqMTMMlOyCIW/DIUk0sNxTZOtp2cJlR5WLkIzTnmVx21I4U1jXUw+3KrjImCh3j0lPRV1RDpcRJQfqlyXsRcRGQ9OOssvaO+tIc0T0k6SSOx8pdY/Y7TyZ9o1Ktp6NXesl5tWPEfv8Axj1FUZT1IQRw8kgEtLURE7EL4Tzv6yv8isXUMhABHmZmdyuKUrSsVrGoAABhsAAAAAAANM43MDTOA2ZbiX2YqR+d+tfFmEZ3EvsxUj879a+LMAAAAAAAD4R0ZCwEKuKjYhqHYbK63HFElJfExijuJ9FtuKR9qrXona6YdwyP3cEYtjo6qJqKn5NFPrYlzyyU6sj2GaySavelP7xm0JQVIQ0MTCZFCOFYiNTqTWo+u55gObT9USCfXTKZmzELIrm3mlZFy6KiI7ddh9KgqGTSBlLs3mDUKSvVSdzWr3JK5n9w6mU0BTUqqBM5gYNbTyEmSG98NTaFH+kRHne1y226hgcmlcLWWLs+KeEt+HgVOIbZ0zIjJDhISWXFtOxcZ+8Bl7WKdGLd0Dj30Ff11Qy7fgV/wGXSyYQMzg0RkvimYphexxtVyvydR9Q6aIoakX4c2F0/ApSZWuhvQV/iKx/iPxQ9HQNJlGlBRUS8UU5paLiuChJX0SIuWx5nx9QDG6GqtuVUvNZnVlTNRaYOIQ3ELbJTxQylWIkGaEnpHcyvo3Ir7dtsJwkxdhYeNnp1tUzxsqcb+zt8h1r4N3NK2gg7ZaG3/uORuYpXLpzQ09l80g2ouFcjkGtp0rpVZBGX4kQ6vc7UvT09mlWNziUQsaiFeZSwTqL72Rqevb/CX3CXFKVi0T7KCeTyc98F8cxHVvt3+0+fv28fleZ/PpPIYcn5tHswqVeqSjM1K9ySuZ/AhjH9K1Haej6XE2vbS9GVb38oxKDlkPWGM04hZ4bj8PBIWbbRLNJGlC0pSnLMi4RnlbMUf8i6T3ve/wAnpda1r7yV/v2iGv3NkM+lE9hzflMezFJTbSJJ2Um+y6TzL4kPlUVTSKn0JObTFqHUorpbzUtRcpJK5269gl00ljFFYvSREkUtmGj1NpWzpmZES1mhSc9pbDK/H7goiTwVZ4gVDHz5CopMM8e9smoyTmtRJI7cREm1hkZkxilRjjugqYPNFexKXDrt+BGMvgIyEmEIiLgohqIYcK6HG1EpJ/EdHF0LSMTDqYXIYJCTK2k0jQUXuMrGP1QtKQlJwMRCwsXERBPum4ZunkniIiIsiytc+O3wAZCAAADTONzA0zgNmW4l9mKkfnfrXxZhGdxL7MVI/O/WvizAAAAAAAAxuvaQgKtl7bES4uHfYM1MPoTc0X2kZcZHYssthDC04UztsiQ3WkQlCckkTaysXu08hmOIrVWrlTa6TiG0PIXd5s0p03E5W0TVl7y4+XiPDUV/XsIkmY6i3XHk7Voh3UkfXxl9xgOrmTVR4aT6WRT89emMvilml5CjVomkjLSI0mZ2Ox3Iy5Pv7TC3+tmrv2sR/PHHYk1YV/UMFHVJAfZsqhFXJlaDRcrkZkSVcIzOxEZnlbZyDk1VTdUU5WcRVdJsJi24o1KeYJOkZGrNRGm5GojPPLO/44FYASZeIlcqQbTVEPk/axGcO8ZEf6tiP8R2uFtO1AzOI6pakWtuJiyUTcOa/V0lEZqMiyTsIiLkuMjGNyT+a05/tyf4CHH3LP8Aritf7RD/AMT47rczyGcyCnZqxOpbEwDrsWlaEvI0TUnQIrkMZomEq7DbEyNlh0+9Hymdx7LZxqEKNCGzdMkuaREZEZE4d0nbZyWMS7TFpvEPPYa2xY+Ne0TqN77eNxMQ+qV1QjFeoTpNpDkdpu74SjRbe98Tf1zItuj1jId/xq/4Nj74b/yHIrGk6lldWu1ZRyyccezfh7lpXO2kVjyUk7Xttvs4rccsQa9JHo6qJdOKLK5Qr2jf9X/2Ib0LkUpR9UzGsWKnrJ1tLkMRG0ylSTMzK+iXA4JJIzvtuZ/EfHA/86Kq/bJ/jcHPomW15MalKf1JHPQEMlNigkKsThcRGgjsRFc8z4Q6udyCrKPq+NqGloX0+DjFKU4wSTWZaR6RpNBWM872NP8A9CuAJMvEKunUGzD0S+l8ysSjh3jIj5bWL947rCim55Ln4+d1C6soyOPJg16WgRq0lGdjsRmdsuK3WMjPwAAAaZxuYGmcBsy3EvsxUj879a+LMIzuJfZipH53618WYAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABpnG5gaZwFzwt3UFfYd0JLqOkkopmIgJfvu9ORkM+p1W+OrdPSNLyS9ZZ2sRZW94ybXVxT5gozscT5gAANdXFPmCjOxxPmA11cU+YKM7HE+YAADXVxT5gozscT5gNdXFPmCjOxxPmAAA11cU+YKM7HE+YDXVxT5gozscT5gAANdXFPmCjOxxPmA11cU+YKM7HE+YAADXVxT5gozscT5gNdXFPmCjOxxPmAAA11cU+YKM7HE+YDXVxT5gozscT5gAANdXFPmCjOxxPmA11cU+YKM7HE+YAADXVxT5gozscT5gNdXFPmCjOxxPmAAA11cU+YKM7HE+YDXVxT5gozscT5gAANdXFPmCjOxxPmA11cU+YKM7HE+YAADXVxT5gozscT5geZgAB/9k="

LOGO_OCP_URI = _logo_b64()

# ─────────────────────────────────────────────────────────────────────
#  STREAMLIT CONFIG
# ─────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Scitec | Prospecção Inteligente", page_icon="🔬",
                   layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=Space+Mono:wght@400;700&display=swap');

/* ── Theme tokens: overridden per module via JS ── */
:root {
  --acc:    #0D9291;
  --acc-dk: #0A7170;
  --acc-lt: #E0F4F4;
  --acc-vl: #7C3AED;
  --mod-badge-bg: #E0F4F4;
  --mod-badge-cl: #0A7170;
}
:root[data-module="ocp"] {
  --acc:    #7C1934;
  --acc-dk: #621427;
  --acc-lt: #F5E0E5;
  --mod-badge-bg: #F5E0E5;
  --mod-badge-cl: #621427;
}

html,body,.stApp{background-color:#FFFFFF!important;color:#4A4A4A;font-family:'DM Sans',sans-serif;}
section[data-testid="stSidebar"]{background-color:#F8F9FA!important;border-right:1px solid #E0E0E0;}
section[data-testid="stSidebar"] *{color:#4A4A4A!important;}

.stSelectbox>div>div,.stTextInput>div>div>input,.stTextArea>div>div>textarea{
    background:#F1F3F5!important;border:1.5px solid #E0E0E0!important;
    border-radius:8px!important;color:#4A4A4A!important;}
.stSelectbox>div>div:focus-within,.stTextInput>div>div>input:focus{
    border-color:var(--acc)!important;box-shadow:0 0 0 3px rgba(13,146,145,.12)!important;}

.stTabs [data-baseweb="tab-list"]{gap:4px;background:#F1F3F5;border-radius:12px;padding:4px;border:1px solid #E0E0E0;}
.stTabs [data-baseweb="tab"]{background:transparent;border-radius:8px;padding:9px 16px;color:#888;
    font-weight:500;font-size:.86rem;border:none!important;transition:all .2s;}
.stTabs [aria-selected="true"]{background-color:var(--acc)!important;color:white!important;
    box-shadow:0 2px 8px rgba(13,146,145,.20);}

.stButton>button{
    background-color:var(--acc)!important;color:white!important;
    border:none!important;border-radius:8px!important;
    padding:.6rem 1.4rem!important;font-size:.93rem!important;
    font-weight:600!important;transition:all .2s!important;width:100%;}
.stButton>button:hover{background-color:var(--acc-dk)!important;
    box-shadow:0 4px 14px rgba(0,0,0,.18)!important;transform:translateY(-1px);}
.stButton>button:active{transform:translateY(0)!important;}
.stDownloadButton>button{background:transparent!important;color:var(--acc)!important;
    border:2px solid var(--acc)!important;border-radius:8px!important;
    padding:.55rem 1.2rem!important;font-weight:600!important;width:100%;}
.stDownloadButton>button:hover{background:var(--acc-lt)!important;}

.stDataFrame{border:1px solid #E0E0E0!important;border-radius:12px!important;overflow:hidden;}
.stProgress>div>div{background:var(--acc)!important;border-radius:4px;}
.stExpander{border:1px solid #E0E0E0!important;border-radius:12px!important;background:#FFF!important;}
.stExpander:hover{border-color:var(--acc)!important;}
.stAlert{border-radius:10px!important;border-left:4px solid var(--acc)!important;}

/* ── Header ── */
.header-bar{background:#FFF;border-bottom:3px solid var(--acc);padding:.8rem 1.5rem;
    margin:-1rem -1rem 1.5rem -1rem;display:flex;justify-content:space-between;
    align-items:center;box-shadow:0 2px 10px rgba(0,0,0,.05);}
.logo-text{font-family:'Space Mono',monospace;font-size:1.3rem;font-weight:700;
    color:var(--acc);letter-spacing:-.02em;}
.hdr-sub{font-size:.72rem;color:#888;margin-top:1px;letter-spacing:.05em;}

/* ── Badges ── */
.badge{padding:4px 11px;border-radius:20px;font-size:11px;font-weight:700;
    font-family:'Space Mono',monospace;color:#fff;display:inline-block;}
.b-acc{background:var(--acc);}.b-hot{background:#E53935;}.b-mu{background:#888;}
.b-am{background:#F59E0B;}.b-bl{background:#3B82F6;}.b-vl{background:#7C3AED;}.b-endo{background:#7C3AED;}
.b-ocp{background:#7C1934;}

/* ── Sections ── */
.sec-title{font-size:1.1rem;font-weight:700;color:#2D2D2D;margin-bottom:.3rem;
    border-left:4px solid var(--acc);padding-left:10px;}
.sec-desc{font-size:.84rem;color:#888;margin-bottom:1rem;padding-left:14px;}

/* ── Metric cards ── */
.mc{background:#FFF;border:1.5px solid #E0E0E0;border-radius:14px;padding:1.1rem 1rem;
    text-align:center;box-shadow:0 2px 8px rgba(0,0,0,.04);
    transition:all .2s;min-height:95px;display:flex;flex-direction:column;justify-content:center;
    cursor:pointer;}
.mc:hover{box-shadow:0 6px 20px rgba(0,0,0,.12);border-color:var(--acc);transform:translateY(-2px);}
.mc.active{border-color:var(--acc);border-width:2px;background:var(--acc-lt);}
.ml{font-size:.7rem;font-weight:600;color:#888;letter-spacing:.07em;text-transform:uppercase;margin-bottom:.3rem;}
.mn{font-size:2rem;font-weight:700;color:var(--acc);font-family:'Space Mono',monospace;line-height:1;}
.mn-hot{color:#E53935;}.mn-am{color:#F59E0B;}.mn-bl{color:#3B82F6;}.mn-vl{color:#7C3AED;}

/* ── Robot CTA ── */
.robot-cta{background:var(--acc-lt);border:1.5px dashed var(--acc);border-radius:14px;padding:1.2rem;
    margin-bottom:.9rem;display:flex;flex-direction:column;align-items:center;gap:.35rem;text-align:center;}
.rc-title{font-size:.95rem;font-weight:600;color:var(--acc-dk);}
.rc-sub{font-size:.78rem;color:#888;}

/* ── Info rows ── */
.ir{display:flex;gap:.4rem;font-size:.84rem;margin-top:.3rem;}
.il{font-weight:600;color:#888;min-width:110px;}.iv{color:#4A4A4A;}

/* ── Empty state ── */
.empty{text-align:center;padding:2.5rem 1rem;color:#888;}
.empty-i{font-size:2.2rem;margin-bottom:.45rem;}.empty-t{font-size:.9rem;font-weight:500;}

/* ── Chips ── */
.pchip{display:inline-block;background:var(--acc-lt);color:var(--acc-dk);font-family:'Space Mono',monospace;
    font-size:.74rem;font-weight:700;padding:3px 10px;border-radius:6px;}

/* ── Banners ── */
.inm-banner{background:var(--acc-lt);border-radius:10px;padding:.7rem 1rem;font-size:.78rem;}
.inm-banner-title{font-weight:700;color:var(--acc-dk);margin-bottom:2px;}
.inm-banner-sub{color:var(--acc);font-size:.74rem;}
.endo-banner{background:#EDE9FE;border-radius:10px;padding:.7rem 1rem;font-size:.78rem;border-left:3px solid #7C3AED;}
.endo-banner-title{font-weight:700;color:#5B21B6;margin-bottom:2px;}
.endo-banner-sub{color:#6D28D9;font-size:.74rem;}
.ocp-banner{background:#F5E0E5;border-radius:10px;padding:.7rem 1rem;font-size:.78rem;border-left:3px solid #7C1934;}
.ocp-banner-title{font-weight:700;color:#621427;margin-bottom:2px;}
.ocp-banner-sub{color:#7C1934;font-size:.74rem;}

/* ── Progress bars ── */
.progress-bar-wrap{background:#F1F3F5;border-radius:6px;height:8px;overflow:hidden;margin-top:4px;}
.progress-bar{height:8px;border-radius:6px;transition:width .5s;}

/* ── Filter pill ── */
.filter-active{background:var(--acc-lt);border:1px solid var(--acc);border-radius:6px;
    padding:4px 10px;font-size:.8rem;font-weight:600;color:var(--acc-dk);display:inline-block;margin-bottom:.5rem;}

/* ── ICP / Persona ── */
.icp-tag{display:inline-block;background:#EDE9FE;color:#5B21B6;border-radius:6px;padding:3px 9px;font-size:.74rem;font-weight:600;margin:2px;}
.icp-ref{background:var(--acc-lt);color:var(--acc-dk);border-radius:6px;padding:3px 9px;font-size:.74rem;margin:2px;display:inline-block;}
.enrich-card{background:#FFF;border:1.5px solid #E0E0E0;border-radius:10px;padding:.7rem 1rem;margin-bottom:.4rem;transition:border-color .2s;}
.enrich-card:hover{border-color:var(--acc);}
.enrich-card.found{border-color:var(--acc);background:var(--acc-lt);}

/* ── Module selector sidebar ── */
.mod-selector{display:flex;gap:6px;margin-bottom:.8rem;}
.mod-btn{flex:1;padding:.5rem;border-radius:8px;border:1.5px solid #E0E0E0;background:#FFF;
    font-size:.8rem;font-weight:600;cursor:pointer;text-align:center;transition:all .2s;color:#888;}
.mod-btn.active-lab{background:#E0F4F4;border-color:#0D9291;color:#0A7170;}
.mod-btn.active-ocp{background:#F5E0E5;border-color:#7C1934;color:#621427;}
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────
#  INSTÂNCIAS
# ─────────────────────────────────────────────────────────────────────
db       = Database("prospec_inmetro.db")
scraper  = InmetroScraper()
enricher = DataEnricher()
exporter = DataExporter(db)
endo_bot = EndotoxinaBot()
crm_rdr = CRMAnalyzer()

# ─────────────────────────────────────────────────────────────────────
#  SESSION STATE (filtro ativo do dashboard)
# ─────────────────────────────────────────────────────────────────────
if "dash_filter"          not in st.session_state: st.session_state["dash_filter"]          = None
if "dash_leads"           not in st.session_state: st.session_state["dash_leads"]           = None
if "persona_cnae_results" not in st.session_state: st.session_state["persona_cnae_results"] = []
if "persona_sem_results"  not in st.session_state: st.session_state["persona_sem_results"]  = []

# ─────────────────────────────────────────────────────────────────────
#  SIDEBAR
# ─────────────────────────────────────────────────────────────────────
with st.sidebar:
    # ── Logo ─────────────────────────────────────────────────────────
    if LOGO_OCP_URI:
        st.markdown(
            f'<div style="text-align:center;padding:.6rem 0 .3rem;">' +
            f'<img src="{LOGO_OCP_URI}" style="max-width:160px;width:100%;height:auto;" alt="Scitec"></div>',
            unsafe_allow_html=True
        )
    st.markdown("<hr style='border:1px solid #E0E0E0;margin:.4rem 0;'>", unsafe_allow_html=True)

    # ── Seletor de Módulo ─────────────────────────────────────────────
    st.markdown('<p style="font-size:.7rem;font-weight:700;letter-spacing:.07em;text-transform:uppercase;color:#888;margin-bottom:6px;">🏢 Unidade</p>', unsafe_allow_html=True)
    modulo = st.radio("Módulo", ["🔬 Laboratório", "🏅 Certificadora (OCP)"],
                      label_visibility="collapsed", horizontal=True, key="modulo_sel")
    IS_LAB = modulo == "🔬 Laboratório"
    IS_OCP = not IS_LAB

    # Inject CSS theme based on module
    _theme_class = "lab" if IS_LAB else "ocp"
    _acc_color   = "#0D9291" if IS_LAB else "#7C1934"
    _acc_dk      = "#0A7170" if IS_LAB else "#621427"
    _acc_lt      = "#E0F4F4" if IS_LAB else "#F5E0E5"
    st.markdown(
        f'<script>document.documentElement.setAttribute("data-module","{_theme_class}")</script>' +
        f'<style>:root{{--acc:{_acc_color};--acc-dk:{_acc_dk};--acc-lt:{_acc_lt};}}</style>',
        unsafe_allow_html=True
    )

    st.markdown("<hr style='border:1px solid #E0E0E0;margin:.5rem 0;'>", unsafe_allow_html=True)

    # ── Menu por módulo ───────────────────────────────────────────────
    if IS_LAB:
        st.markdown('<p style="font-size:.7rem;font-weight:700;letter-spacing:.07em;text-transform:uppercase;color:#888;margin-bottom:4px;">🔬 Ensaios — Laboratório</p>', unsafe_allow_html=True)
        _LAB_MENUS = {
            "🧬 Endotoxina & Esterilidade":   "ENDO_EST",
            "🦴 ISO 10993-18 — Biocompatib.": "ISO10993",
            "🧲 MRI — Ressonância Magnética":  "MRI",
            "⚗️ Bioburden (CNAE 3250-7/01)":  "BIOBURDEN",
        }
        menu_label = st.selectbox("Ensaio", list(_LAB_MENUS.keys()), label_visibility="collapsed", key="menu_lab")
        pid_cod = _LAB_MENUS[menu_label]
        portaria_id = pid_cod
        IS_ENDO_SCOPE = True
        estado, cidade = "Todos", ""
    else:
        st.markdown('<p style="font-size:.7rem;font-weight:700;letter-spacing:.07em;text-transform:uppercase;color:#888;margin-bottom:4px;">🏅 Portaria — OCP</p>', unsafe_allow_html=True)
        _OCP_MENUS = {
            "Portaria 145/2022 — Automotivos":    "145",
            "Portaria 384/2020 — Eletromédicos":  "384",
        }
        menu_label = st.selectbox("Portaria", list(_OCP_MENUS.keys()), label_visibility="collapsed", key="menu_ocp")
        pid_cod = _OCP_MENUS[menu_label]
        portaria_id = pid_cod
        IS_ENDO_SCOPE = False
        st.markdown("<hr style='border:1px solid #E0E0E0;margin:.5rem 0;'>", unsafe_allow_html=True)
        st.markdown('<p style="font-size:.7rem;font-weight:700;letter-spacing:.07em;text-transform:uppercase;color:#888;margin-bottom:4px;">🌎 Filtros Geográficos</p>', unsafe_allow_html=True)
        estado = st.selectbox("Estado", ["Todos","SP","RJ","MG","PR","SC","RS","GO","BA","CE","PE","DF","MT","MS","AM","PA"], key="estado_ocp")
        cidade = st.text_input("Cidade", placeholder="Ex: São Paulo", key="cidade_ocp")

    st.markdown("<hr style='border:1px solid #E0E0E0;margin:.5rem 0;'>", unsafe_allow_html=True)

    # ── Banner de contexto ────────────────────────────────────────────
    if IS_LAB:
        _LAB_SCOPE_INFO = {
            "ENDO_EST":  ("🧬","Endotoxina & Esterilidade","CNAE 8129-0/00 · 8650-0/99"),
            "ISO10993":  ("🦴","ISO 10993-18","CNAE 3250-7/01 · Dispositivos implantáveis"),
            "MRI":       ("🧲","MRI","CNAE 26.60-4/00 · Equip. de imagem"),
            "BIOBURDEN": ("⚗️","Bioburden","CNAE 3250-7/01 · Fab. equip. médicos"),
        }
        ico_sb, tit_sb, sub_sb = _LAB_SCOPE_INFO.get(pid_cod, ("🔬","Laboratório",""))
        st.markdown(
            f'<div class="endo-banner"><div class="endo-banner-title">{ico_sb} {tit_sb}</div>' +
            f'<div class="endo-banner-sub">{sub_sb}</div></div>',
            unsafe_allow_html=True
        )
    else:
        st.markdown(
            f'<div class="ocp-banner"><div class="ocp-banner-title">🏅 Portaria Ativa</div>' +
            f'<div class="ocp-banner-sub"><strong>Nº {pid_cod}</strong> · INMETRO + BrasilAPI</div></div>',
            unsafe_allow_html=True
        )

    st.markdown("<hr style='border:1px solid #E0E0E0;margin:.5rem 0;'>", unsafe_allow_html=True)
    st.markdown('<p style="font-size:.7rem;font-weight:700;letter-spacing:.07em;text-transform:uppercase;color:#888;margin-bottom:4px;">📂 CRM — RD Station</p>', unsafe_allow_html=True)
    crm_files = st.file_uploader("CSV RD Station", type=["csv"], accept_multiple_files=True, label_visibility="collapsed", key="crm_uploader")

# ── CRM ──────────────────────────────────────────────────────────────
crm_slugs, crm_df_all = set(), None
if crm_files:
    dfs = []
    for f in crm_files:
        try: dfs.append(crm_rdr.load_csv(f.read()))
        except: pass
    if dfs:
        crm_df_all = pd.concat(dfs, ignore_index=True)
        crm_slugs  = crm_rdr.empresa_slugs(crm_df_all)

# ── Header ─────────────────────────────────────────────────────────────────
badge_crm = (f'<span class="badge b-bl">CRM: {len(crm_df_all)}</span>'
             if crm_df_all is not None else "")

if IS_LAB:
    _SCOPE_BADGES = {
        "ENDO_EST":  '<span class="badge b-vl">🧬 Endotoxina & Esterilidade</span>',
        "ISO10993":  '<span class="badge b-vl">🦴 ISO 10993-18</span>',
        "MRI":       '<span class="badge b-vl">🧲 MRI</span>',
        "BIOBURDEN": '<span class="badge b-acc">⚗️ Bioburden</span>',
    }
    badge_modo = _SCOPE_BADGES.get(pid_cod, '<span class="badge b-acc">🔬 Laboratório</span>')
    _sub_txt  = "LABORATÓRIO · ENSAIOS BIOLÓGICOS"
else:
    badge_modo = f'<span class="badge b-ocp">🏅 OCP · Portaria {pid_cod}</span><span class="badge b-mu">INMETRO</span>'
    _sub_txt   = "CERTIFICADORA OCP · INMETRO"

_logo_img = (f'<img src="{LOGO_OCP_URI}" style="height:44px;width:auto;" alt="Scitec">'
             if LOGO_OCP_URI else "")
_border_color = _acc_color if "IS_LAB" else "#7C1934"
st.markdown(
    f'<div class="header-bar" style="border-bottom-color:{_acc_color};">' +
    f'<div style="display:flex;align-items:center;gap:14px;">{_logo_img}' +
    f'<div><div class="logo-text" style="color:{_acc_color};">PROSPEC-O</div>' +
    f'<div class="hdr-sub">{_sub_txt} &nbsp;·&nbsp; SCITEC CERTIFICAÇÕES</div></div></div>' +
    f'<div style="display:flex;gap:7px;align-items:center;flex-wrap:wrap;">{badge_modo}{badge_crm}</div>' +
    '</div>',
    unsafe_allow_html=True
)
# ─────────────────────────────────────────────────────────────────────
#  ABAS
# ─────────────────────────────────────────────────────────────────────
# Nomes de abas adaptados ao módulo
_LAB_TABS = ["🔍 Buscar Leads","📋 Leads","🔵 Funil CRM","🔥 Oportunidades","🧬 Ensaios","🎯 Persona ICP","📊 Painel"]
_OCP_TABS = ["🔍 Buscar Leads","📋 Leads","🔵 Funil CRM","🔥 Oportunidades","🧬 Histórico","🎯 Persona ICP","📊 Painel"]
t1,t2,t3,t4,t5,t6,t7 = st.tabs(_LAB_TABS if IS_LAB else _OCP_TABS)

# ══════════════════════════════════════════════════════════════════════
#  ABA 1 — COLETA
# ══════════════════════════════════════════════════════════════════════
with t1:
    iniciar = limpar = False
    if IS_LAB:
        st.markdown(
            f'<div style="background:{_acc_lt};border:1.5px dashed {_acc_color};border-radius:14px;padding:1.5rem;text-align:center;margin-top:.5rem;">' +
            f'<div style="font-size:1.8rem;margin-bottom:.4rem;">🔬</div>' +
            f'<div style="font-weight:700;color:{_acc_dk};margin-bottom:.3rem;">Módulo Laboratório — {menu_label}</div>' +
            f'<div style="font-size:.82rem;color:{_acc_color};">Acesse a aba <strong>🧬 Endotoxina</strong> para prospectar clientes por CNAE.</div></div>',
            unsafe_allow_html=True
        )
    else:
        st.markdown(
            f'<div class="sec-title" style="border-left-color:#7C1934;">🏅 Buscar Leads — Portaria {pid_cod}</div>',
            unsafe_allow_html=True
        )
        st.markdown('<div class="sec-desc">Busca via dados.gov.br → HTML INMETRO (paginado, até 20 páginas) → CSV direto → BrasilAPI + cruzamento ICP.</div>', unsafe_allow_html=True)

        col_btn, col_info = st.columns([1,3], gap="large")
        with col_btn:
            st.markdown('<div class="robot-cta"><div style="font-size:1.7rem;">🤖</div><div class="rc-title">Buscar Leads</div><div class="rc-sub">CNAE · INMETRO · BrasilAPI · CRM ICP</div></div>', unsafe_allow_html=True)
            iniciar   = st.button("🔍  Buscar Leads", key="btn_robot", use_container_width=True)
            st.markdown("<div style='margin-top:5px;'></div>", unsafe_allow_html=True)
            limpar    = st.button("🗑  Limpar Base",  key="btn_clear", use_container_width=True)

        with col_info:
            st.markdown("""
            <div style="background:#FFF;border:1.5px solid #E0E0E0;border-radius:14px;padding:1.2rem 1.4rem;">
              <div style="font-weight:700;font-size:.9rem;color:#2D2D2D;margin-bottom:.8rem;">🔍 Como a busca funciona</div>
              <div style="display:flex;flex-direction:column;gap:.6rem;">
                <div style="display:flex;gap:10px;align-items:flex-start;">
                  <span style="background:#E0F4F4;border-radius:50%;width:24px;height:24px;min-width:24px;display:flex;align-items:center;justify-content:center;font-size:.7rem;font-weight:700;color:#0A7170;">1</span>
                  <div><strong>dados.gov.br CKAN</strong><br><span style="font-size:.78rem;color:#888;">API pública — datasets INMETRO no portal de dados abertos do governo federal.</span></div>
                </div>
                <div style="display:flex;gap:10px;align-items:flex-start;">
                  <span style="background:#E0F4F4;border-radius:50%;width:24px;height:24px;min-width:24px;display:flex;align-items:center;justify-content:center;font-size:.7rem;font-weight:700;color:#0A7170;">2</span>
                  <div><strong>Scraping HTML com paginação</strong><br><span style="font-size:.78rem;color:#888;">Extrai até 20 páginas de consulta.inmetro.gov.br automaticamente.</span></div>
                </div>
                <div style="display:flex;gap:10px;align-items:flex-start;">
                  <span style="background:#E0F4F4;border-radius:50%;width:24px;height:24px;min-width:24px;display:flex;align-items:center;justify-content:center;font-size:.7rem;font-weight:700;color:#0A7170;">3</span>
                  <div><strong>CSV/XLSX direto INMETRO</strong><br><span style="font-size:.78rem;color:#888;">Download de listas OCP e produtos certificados publicados pelo INMETRO.</span></div>
                </div>
                <div style="display:flex;gap:10px;align-items:flex-start;">
                  <span style="background:#E0F4F4;border-radius:50%;width:24px;height:24px;min-width:24px;display:flex;align-items:center;justify-content:center;font-size:.7rem;font-weight:700;color:#0A7170;">4</span>
                  <div><strong>BrasilAPI + Receita Federal</strong><br><span style="font-size:.78rem;color:#888;">CNPJ completo, QSA (decisor), setor CNAE, telefone — sem chave de API.</span></div>
                </div>
              </div>
              <div style="margin-top:.8rem;padding:.55rem .85rem;background:#FFF8E1;border-left:3px solid #F59E0B;border-radius:6px;font-size:.78rem;color:#92400E;">
                <strong>Sem acesso externo?</strong> A busca usa APIs públicas (dados.gov.br, Casa dos Dados). Verifique sua conexão.
              </div>
            </div>""", unsafe_allow_html=True)

        if limpar:
            db.clear_leads(pid_cod)
            st.warning(f"Base da portaria **{pid_cod}** limpa."); st.rerun()



        if iniciar:
            with st.status("🤖 Prospecção em andamento...", expanded=True) as sc:
                st.write("🔗 Fonte 1 — dados.gov.br...")
                leads_raw = scraper.fetch_all_empresas(
                    portaria=pid_cod,
                    estado=estado if estado!="Todos" else None,
                    cidade=cidade or None)
                if not leads_raw:
                    sc.update(label="⚠️ Fontes externas indisponíveis.", state="error")
                    st.warning("Portal INMETRO inacessível. Verifique sua conexão de rede.")
                else:
                    fonte = leads_raw[0].get("fonte","inmetro")
                    st.write(f"✅ {len(leads_raw)} registros via **{fonte}**. Enriquecendo com BrasilAPI...")
                    prog2 = st.progress(0)
                    for i,l in enumerate(leads_raw):
                        l = enricher.enrich_receita(l)
                        l = enricher.enrich_contatos(l)
                        db.upsert_lead(l, pid_cod)
                        prog2.progress((i+1)/len(leads_raw))
                    sc.update(label=f"✅ {len(leads_raw)} empresas coletadas • Fonte: {fonte}", state="complete")
                    st.balloons()
                    st.success(f"**{len(leads_raw)} empresas** salvas. Acesse **📋 Leads**.")

# ══════════════════════════════════════════════════════════════════════
#  ABA 2 — LEADS (com busca e filtros avançados)
# ══════════════════════════════════════════════════════════════════════
with t2:
    if IS_LAB:
        st.markdown(
            f'<div style="background:{_acc_lt};border:1.5px dashed {_acc_color};border-radius:14px;padding:1.5rem;text-align:center;">' +
            f'<div style="font-size:1.8rem;">🔬</div>' +
            f'<div style="font-weight:700;color:{_acc_dk};margin:.3rem 0;">Módulo Laboratório — {menu_label}</div>' +
            f'<div style="font-size:.82rem;color:{_acc_color};">Acesse a aba <strong>🧬 Endotoxina</strong>.</div></div>',
            unsafe_allow_html=True
        )
    else:
        st.markdown(f'<div class="sec-title">Base de Prospecção — Portaria {pid_cod}</div>', unsafe_allow_html=True)
        st.markdown('<div class="sec-desc">Marque ✔️ para iniciar régua de 6 meses · ✖️ para excluir da visão ativa.</div>', unsafe_allow_html=True)

        # Filtros avançados
        with st.expander("🔍 Filtros avançados", expanded=False):
            fa1,fa2,fa3 = st.columns(3, gap="small")
            with fa1: busca_texto = st.text_input("Buscar por nome / CNPJ / cidade", key="busca_leads")
            with fa2:
                ocps_disponiveis = ["Todos"] + sorted(set(l.get("ocp","") or "" for l in db.get_leads({"portaria":pid_cod}) if l.get("ocp")))
                filtro_ocp = st.selectbox("Filtrar por OCP", ocps_disponiveis, key="filtro_ocp")
            with fa3:
                filtro_urg = st.selectbox("Urgência", ["Todos","hot","medio","normal","vencido"], key="filtro_urg")

        filtros = {"portaria":pid_cod,"estado":estado,"cidade":cidade}
        if busca_texto: filtros["busca"] = busca_texto
        if filtro_ocp and filtro_ocp!="Todos": filtros["ocp"] = filtro_ocp
        if filtro_urg and filtro_urg!="Todos": filtros["urgencia"] = filtro_urg

        leads_db = db.get_leads(filtros)

        if leads_db:
            df = pd.DataFrame(leads_db)
            total    = len(df)
            c_email  = int(df["email"].notna().sum())   if "email"   in df.columns else 0
            c_dec    = int(df["decisor"].notna().sum()) if "decisor" in df.columns else 0
            c_cont   = int((df.get("status_crm","")=="contatado").sum()) if "status_crm" in df.columns else 0
            pct_email  = round(c_email/total*100)   if total else 0
            pct_dec    = round(c_dec/total*100)     if total else 0

            m1,m2,m3,m4 = st.columns(4, gap="small")
            with m1: st.markdown(f'<div class="mc"><div class="ml">Total de Leads</div><div class="mn">{total}</div></div>', unsafe_allow_html=True)
            with m2:
                st.markdown(f'<div class="mc"><div class="ml">Com E-mail</div><div class="mn">{c_email}</div><div class="progress-bar-wrap"><div class="progress-bar" style="width:{pct_email}%;background:#0D9291;"></div></div></div>', unsafe_allow_html=True)
            with m3:
                st.markdown(f'<div class="mc"><div class="ml">Decisor ID</div><div class="mn">{c_dec}</div><div class="progress-bar-wrap"><div class="progress-bar" style="width:{pct_dec}%;background:#3B82F6;"></div></div></div>', unsafe_allow_html=True)
            with m4: st.markdown(f'<div class="mc"><div class="ml">Contatados ✔️</div><div class="mn mn-am">{c_cont}</div></div>', unsafe_allow_html=True)

            if busca_texto or (filtro_ocp and filtro_ocp!="Todos") or (filtro_urg and filtro_urg!="Todos"):
                ativos = []
                if busca_texto: ativos.append(f"Busca: {busca_texto}")
                if filtro_ocp and filtro_ocp!="Todos": ativos.append(f"OCP: {filtro_ocp}")
                if filtro_urg and filtro_urg!="Todos": ativos.append(f"Urgência: {filtro_urg}")
                st.markdown(f'<div class="filter-active">🔽 Filtro ativo: {" · ".join(ativos)}</div>', unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)
            urgencia_color = {"hot":"🔥","vencido":"⚠️","medio":"⏳","normal":"✅"}
            df["🏷"]  = df["urgencia"].map(urgencia_color).fillna("—")
            # CRM status merge em tempo real
            if crm_df_all is not None:
                df["CRM"] = df.apply(
                    lambda row: CRMAnalyzer.crm_status(
                        str(row.get("nome","") or ""),
                        crm_df_all,
                        str(row.get("cnpj","") or "")
                    ), axis=1)
                crm_icon = {"vendida":"✅","perdida":"❌","em_andamento":"🔄","novo":"🆕"}
                df["CRM"] = df["CRM"].map(lambda x: crm_icon.get(x,"🆕")+" "+x.title().replace("_"," "))
                # Enriquece email/telefone/decisor do CRM quando vazios
                def _fill_from_crm(row):
                    detail = CRMAnalyzer.crm_detail(
                        str(row.get("nome","") or ""), crm_df_all,
                        str(row.get("cnpj","") or ""))
                    if detail.get("crm_email") and not row.get("email"):
                        row["email"] = detail["crm_email"]
                    if detail.get("crm_telefone") and not row.get("telefone"):
                        row["telefone"] = detail["crm_telefone"]
                    if detail.get("crm_contato") and not row.get("decisor"):
                        row["decisor"] = detail["crm_contato"]
                    if detail.get("crm_cargo") and not row.get("cargo_decisor"):
                        row["cargo_decisor"] = detail["crm_cargo"]
                    return row
                df = df.apply(_fill_from_crm, axis=1)
            # Lookalike score
            if "lookalike" in df.columns:
                df["Lookalike"] = df["lookalike"].apply(lambda s: "🔥" if (s or 0)>=70 else ("⭐" if (s or 0)>=40 else ("➕" if (s or 0)>0 else "—")))
            cols_v = ["🏷","nome","cnpj","email","telefone","decisor","ocp","validade","estado","cidade","status_crm","setor"]
            if "CRM"       in df.columns: cols_v.append("CRM")
            if "Lookalike" in df.columns: cols_v.append("Lookalike")
            st.dataframe(df[[c for c in cols_v if c in df.columns]], use_container_width=True, height=360)

            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown('<div style="font-weight:600;font-size:.88rem;color:#2D2D2D;margin-bottom:.5rem;">⚙️ Atualizar status do lead</div>', unsafe_allow_html=True)
            cs1,cs2,cs3 = st.columns([3,1,1], gap="small")
            with cs1:
                nomes = [l["nome"] for l in leads_db if l.get("nome")]
                nome_sel = st.selectbox("Lead", nomes, label_visibility="collapsed", key="sel_lead")
            with cs2:
                if st.button("✔️ Contatado (6m)", key="btn_cont", use_container_width=True):
                    sel = next((l for l in leads_db if l["nome"]==nome_sel),None)
                    if sel: db.set_lead_status(sel["id"],"contatado"); st.success(f"✔️ **{nome_sel}** → régua 180 dias."); st.rerun()
            with cs3:
                if st.button("✖️ Não Prospectar", key="btn_np", use_container_width=True):
                    sel = next((l for l in leads_db if l["nome"]==nome_sel),None)
                    if sel: db.set_lead_status(sel["id"],"nao_prospectar"); st.warning(f"✖️ **{nome_sel}** removido."); st.rerun()
        else:
            st.markdown('<div class="empty"><div class="empty-i">📭</div><div class="empty-t">Nenhum lead encontrado.</div><div style="font-size:.78rem;color:#888;margin-top:4px;">Use <strong>🔍 Buscar Leads</strong> para iniciar a prospecção.</div></div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════
#  ABA 3 — FUNIL CRM
# ══════════════════════════════════════════════════════════════════════
with t3:
    if IS_LAB:
        st.markdown(f'<div class="empty"><div class="empty-i">🧬</div><div class="empty-t">Funil CRM disponível no módulo OCP (portarias INMETRO).</div><div style="font-size:.78rem;color:#888;margin-top:4px;">Selecione <strong>🏅 Certificadora (OCP)</strong> no menu lateral.</div></div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="sec-title">Funil CRM — Leads Inéditos</div>', unsafe_allow_html=True)
        st.markdown('<div class="sec-desc">Empresas do INMETRO que <strong>não constam</strong> no seu CRM — prioridade máxima de abordagem.</div>', unsafe_allow_html=True)
        if crm_df_all is None:
            st.markdown('<div style="background:#EFF6FF;border:1.5px dashed #3B82F6;border-radius:14px;padding:2rem;text-align:center;"><div style="font-size:1.8rem;margin-bottom:.4rem;">📂</div><div style="font-weight:600;color:#1E40AF;margin-bottom:.3rem;">Nenhum CSV do RD Station importado</div><div style="font-size:.82rem;color:#6B7280;">Faça o upload na <strong>sidebar</strong> (seção CRM — RD Station).</div></div>', unsafe_allow_html=True)
        else:
            crm_sum = crm_rdr.summary(crm_df_all)
            st.markdown(f'<span class="badge b-bl" style="margin-bottom:.7rem;display:inline-block;">📊 {crm_sum["total"]} negócios no CRM</span>', unsafe_allow_html=True)
            all_leads = db.get_leads({"portaria":pid_cod})
            novos  = [l for l in all_leads
                       if CRMAnalyzer.crm_status(l.get("nome",""), crm_df_all, l.get("cnpj","")) == "novo"]
            ja_crm = [l for l in all_leads
                       if CRMAnalyzer.crm_status(l.get("nome",""), crm_df_all, l.get("cnpj","")) != "novo"]
            ci1,ci2 = st.columns(2, gap="medium")
            with ci1:
                st.markdown("**Etapas no CRM:**")
                for etapa,n in list(crm_sum["etapas"].items())[:8]:
                    pct = round(n/crm_sum["total"]*100) if crm_sum["total"] else 0
                    st.markdown(f'<div style="display:flex;justify-content:space-between;font-size:.82rem;padding:4px 0;border-bottom:1px solid #F1F3F5;"><span>{etapa}</span><span style="font-weight:600;color:#0D9291;">{n} ({pct}%)</span></div>', unsafe_allow_html=True)
            with ci2:
                st.markdown(f'<div style="display:flex;flex-direction:column;gap:.5rem;"><div class="mc"><div class="ml">Leads Inéditos 🆕</div><div class="mn mn-bl">{len(novos)}</div></div><div class="mc"><div class="ml">Já no CRM</div><div class="mn mn-am">{len(ja_crm)}</div></div></div>', unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)

            # ── Análise de performance do CRM ────────────────────────
            if "Estado" in crm_df_all.columns:
                v_count = (crm_df_all["Estado"]=="Vendida").sum()
                p_count = (crm_df_all["Estado"]=="Perdida").sum()
                a_count = (crm_df_all["Estado"]=="Em Andamento").sum()
                taxa_conv = round(v_count/(v_count+p_count)*100) if (v_count+p_count)>0 else 0

                fa,fb,fc,fd = st.columns(4, gap="small")
                with fa: st.markdown(f'<div class="mc"><div class="ml">Negócios CRM</div><div class="mn mn-bl">{crm_sum["total"]}</div></div>', unsafe_allow_html=True)
                with fb: st.markdown(f'<div class="mc"><div class="ml">Ganhos ✅</div><div class="mn" style="color:#16a34a;">{v_count}</div></div>', unsafe_allow_html=True)
                with fc: st.markdown(f'<div class="mc"><div class="ml">Perdidos ❌</div><div class="mn mn-hot">{p_count}</div></div>', unsafe_allow_html=True)
                with fd: st.markdown(f'<div class="mc"><div class="ml">Tx Conversão</div><div class="mn" style="color:#0D9291;">{taxa_conv}%</div></div>', unsafe_allow_html=True)

            # ── Top produtos no CRM ───────────────────────────────────
            if "Produtos" in crm_df_all.columns:
                st.markdown('<div style="font-weight:600;font-size:.88rem;color:#2D2D2D;margin:.8rem 0 .4rem;">📦 Produtos mais solicitados no CRM</div>', unsafe_allow_html=True)
                prods = crm_df_all["Produtos"].value_counts().head(6)
                total_prods = prods.sum()
                for prod, n in prods.items():
                    if not str(prod).strip(): continue
                    pct_p = round(n/total_prods*100)
                    cor_p = "#7C3AED" if "10993" in str(prod) else "#0D9291" if "MRI" in str(prod) else "#3B82F6"
                    st.markdown(f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:4px;"><span style="font-size:.8rem;font-weight:600;min-width:160px;">{prod[:20]}</span><div class="progress-bar-wrap" style="flex:1;"><div class="progress-bar" style="width:{pct_p}%;background:{cor_p};"></div></div><span style="font-size:.78rem;color:#888;min-width:40px;text-align:right;">{n}</span></div>', unsafe_allow_html=True)

            st.markdown("<hr style='border:1px solid #F1F3F5;margin:.8rem 0;'>", unsafe_allow_html=True)

            if novos:
                for lv in novos:
                    lv["_look"] = CRMAnalyzer.lookalike_score(lv)
                novos_sorted = sorted(novos, key=lambda x: x.get("_look",0), reverse=True)
                st.markdown(
                    f'<span class="badge b-ok" style="margin-bottom:.6rem;display:inline-block;">' +
                    f'🆕 {len(novos)} inéditos · ordenados por similaridade ICP</span>',
                    unsafe_allow_html=True
                )
                df_n = pd.DataFrame(novos_sorted)
                df_n["Lookalike"] = df_n["_look"].apply(
                    lambda s: "🔥 Alto" if s>=70 else ("⭐ Médio" if s>=40 else ("➕ Baixo" if s>0 else "—")))
                # Enriquecer com dados do CRM quando disponível
                if crm_df_all is not None:
                    def _enrich_n(row):
                        d = CRMAnalyzer.crm_detail(str(row.get("nome","")), crm_df_all, str(row.get("cnpj","")))
                        if d.get("crm_email")    and not row.get("email"):    row["email"]    = d["crm_email"]
                        if d.get("crm_telefone") and not row.get("telefone"): row["telefone"] = d["crm_telefone"]
                        if d.get("crm_contato")  and not row.get("decisor"):  row["decisor"]  = d["crm_contato"]
                        return row
                    df_n = df_n.apply(_enrich_n, axis=1)
                cols_n = ["nome","cnpj","email","telefone","decisor","Lookalike","ocp","validade","urgencia","estado","cidade","setor"]
                st.dataframe(df_n[[c for c in cols_n if c in df_n.columns]], use_container_width=True, height=360)
                buf_n = io.BytesIO()
                df_n.drop(columns=["_look"], errors="ignore").to_excel(buf_n, index=False)
                st.download_button("⬇  Exportar leads inéditos (.xlsx)", buf_n.getvalue(),
                                   f"leads_ineditos_p{pid_cod}.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            if ja_crm:
                st.markdown("<br>", unsafe_allow_html=True)
                st.markdown(
                    f'<span class="badge b-am" style="margin-bottom:.6rem;display:inline-block;">' +
                    f'📋 {len(ja_crm)} leads já no CRM</span>',
                    unsafe_allow_html=True
                )
                df_ja = pd.DataFrame(ja_crm)
                # Buscar status e produto do CRM
                def _crm_cols(row):
                    d = CRMAnalyzer.crm_detail(str(row.get("nome","")), crm_df_all, str(row.get("cnpj","")))
                    row["CRM Status"] = d.get("crm_estado","—")
                    row["Produto"]    = d.get("crm_produto","—")
                    row["Contato"]    = d.get("crm_contato","—") or row.get("decisor","—")
                    row["Cargo"]      = d.get("crm_cargo","—")
                    return row
                df_ja = df_ja.apply(_crm_cols, axis=1)
                cols_ja = ["nome","cnpj","CRM Status","Produto","Contato","Cargo","email","telefone","estado","cidade"]
                with st.expander("Ver leads já cadastrados no CRM"):
                    st.dataframe(df_ja[[c for c in cols_ja if c in df_ja.columns]], use_container_width=True, height=280)
            if not novos and not ja_crm:
                st.success("Todos os leads do INMETRO já constam no CRM.")

# ══════════════════════════════════════════════════════════════════════
#  ABA 4 — OPORTUNIDADES
# ══════════════════════════════════════════════════════════════════════
with t4:
    if IS_LAB:
        st.markdown(f'<div class="empty"><div class="empty-i">🔔</div><div class="empty-t">Oportunidades de vencimento disponíveis no módulo OCP.</div><div style="font-size:.78rem;color:#888;margin-top:4px;">Selecione <strong>🏅 Certificadora (OCP)</strong>.</div></div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="sec-title">Oportunidades — Vencimentos e Recontatação</div>', unsafe_allow_html=True)
        st.markdown('<div class="sec-desc">Certificações críticas · Leads prontos para recontatação (≥180 dias).</div>', unsafe_allow_html=True)
        hot_leads   = db.get_leads({"portaria":pid_cod,"urgencia":"hot"})
        venc_leads  = db.get_leads({"portaria":pid_cod,"urgencia":"vencido"})
        recon_leads = db.get_recontatar(pid_cod)

        def _opp_card(l, cor):
            icon = {"hot":"🚩","recon":"⏰","vencido":"⚠️"}[cor]
            vc   = {"hot":"#E53935","recon":"#F59E0B","vencido":"#6B7280"}[cor]
            label_val = "Recontatar em" if cor=="recon" else "Validade"
            val_show  = l.get("recontatar_em","—") if cor=="recon" else l.get("validade","—")
            with st.expander(f"{icon}  {l.get('nome','')}  —  {val_show}"):
                c1,c2 = st.columns(2, gap="large")
                with c1: st.markdown(f'<div class="ir"><span class="il">Decisor</span><span class="iv">{l.get("decisor","—")}</span></div><div class="ir"><span class="il">CNPJ</span><span class="iv">{l.get("cnpj","—")}</span></div><div class="ir"><span class="il">OCP</span><span class="iv">{l.get("ocp","—")}</span></div><div class="ir"><span class="il">Setor</span><span class="iv">{l.get("setor","—")}</span></div>', unsafe_allow_html=True)
                with c2: st.markdown(f'<div class="ir"><span class="il">E-mail</span><span class="iv">{l.get("email","—")}</span></div><div class="ir"><span class="il">Telefone</span><span class="iv">{l.get("telefone","—")}</span></div><div class="ir"><span class="il">{label_val}</span><span class="iv" style="color:{vc};font-weight:600;">{val_show}</span></div>', unsafe_allow_html=True)
                if st.button(f"✔️ Marcar como contatado", key=f"opp_cont_{l['id']}", use_container_width=False):
                    db.set_lead_status(l["id"],"contatado"); st.rerun()

        if recon_leads:
            st.markdown(f'<span class="badge b-am" style="margin:.5rem 0;display:inline-block;">⏰ {len(recon_leads)} RECONTATAÇÃO</span>', unsafe_allow_html=True)
            for l in recon_leads: _opp_card(l,"recon")
        if venc_leads:
            st.markdown(f'<span class="badge b-mu" style="margin:.5rem 0;display:inline-block;">⏱ {len(venc_leads)} VENCIDOS</span>', unsafe_allow_html=True)
            for l in venc_leads[:15]: _opp_card(l,"vencido")
        if hot_leads:
            st.markdown(f'<span class="badge b-hot" style="margin:.5rem 0;display:inline-block;">🔥 {len(hot_leads)} QUENTES (≤90 dias)</span>', unsafe_allow_html=True)
            for l in hot_leads: _opp_card(l,"hot")
        if not hot_leads and not venc_leads and not recon_leads:
            st.markdown('<div class="empty"><div class="empty-i">✅</div><div class="empty-t">Nenhuma oportunidade crítica no momento.</div></div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════
#  ABA 5 — ENDOTOXINA
# ══════════════════════════════════════════════════════════════════════
with t5:
    st.markdown('<div class="sec-title">Prospecção — Endotoxina, Esterilidade & Bioburden</div>', unsafe_allow_html=True)
    st.markdown('<div class="sec-desc">Busca <strong>empresas reais</strong> por CNAE — Casa dos Dados · Econodata · DDG empresa-only. Artigos, guias e portais são <em>bloqueados automaticamente</em>.</div>', unsafe_allow_html=True)

    # ── Info de escopo + CNAEs ────────────────────────────────────────
    _ESCOPO_T5 = {
        "ENDO_EST": ("🧬","Endotoxina & Esterilidade",["8129000","8650099","2121101","2122600"],
                     "CNAE 8129-0/00 · 8650-0/99 · 2121-1/01 · 2122-6/00"),
        "ISO10993": ("🦴","ISO 10993 — Biocompatibilidade",["3250701"],
                     "CNAE 3250-7/01 · Fab. equip. e instrumentos médico-odontológicos"),
        "MRI":      ("🧲","MRI — Ressonância Magnética",["2660400"],
                     "CNAE 26.60-4/00 · Fab. equip. irradiação eletromagnética/imagem"),
    }
    escopo_t5 = pid_cod if IS_LAB and pid_cod in _ESCOPO_T5 else "ISO10993"
    ico_t5, titulo_t5, cnaes_t5, desc_t5 = _ESCOPO_T5[escopo_t5]
    st.markdown(
        f'<div style="background:#EDE9FE;border:1.5px solid #7C3AED;border-radius:10px;'
        f'padding:.7rem 1.2rem;margin-bottom:.9rem;">' +
        f'<strong style="color:#5B21B6;">{ico_t5} {titulo_t5}</strong> ' +
        f'<span style="color:#6D28D9;font-size:.82rem;">— {desc_t5}</span>' +
        f'<br><span style="font-size:.76rem;color:#7C3AED;">CNAEs alvo: {" · ".join(cnaes_t5)}</span></div>',
        unsafe_allow_html=True
    )

    col_eb, col_ei = st.columns([1, 3], gap="large")
    with col_eb:
        st.markdown(
            f'<div class="robot-cta">' +
            f'<div style="font-size:1.7rem;">🔍</div>' +
            f'<div class="rc-title">Buscar Leads por CNAE</div>' +
            f'<div class="rc-sub">Casa dos Dados → Econodata → DDG empresa-only</div></div>',
            unsafe_allow_html=True
        )
        fonte_prim   = st.selectbox("Fonte de dados",
                                     ["Casa dos Dados (CNAE)", "Econodata (CNAE)", "DuckDuckGo (fallback)"],
                                     key="fonte_endo_prim")
        ufs_disp     = ["SP","MG","SC","RS","PR","RJ","GO","CE","PE","DF","BA","Todos"]
        uf_endo      = st.multiselect("Estados alvo", ufs_disp,
                                       default=["SP","SC","MG","RS","PR"], key="uf_endo_sel")
        max_pags     = st.slider("Páginas por estado", 1, 5, 2, key="max_pags_endo")
        seeds_raw    = st.text_area(
            "Seeds DDG (fallback)",
            value="\n".join(SEEDS_ISO10993[:3] if escopo_t5=="ISO10993" else
                             SEEDS_MRI[:3]     if escopo_t5=="MRI"       else ENDO_EST_SEEDS[:3]),
            height=90, key="seeds_endo_area"
        )
        seeds_lista  = [s.strip() for s in seeds_raw.splitlines() if s.strip()]
        iniciar_endo = st.button("🔬  Iniciar Busca por CNAE", key="btn_endo", use_container_width=True)
        st.markdown("<div style='margin-top:4px;'></div>", unsafe_allow_html=True)
        if st.button("🗑  Limpar base", key="btn_endo_clear", use_container_width=True):
            db.clear_endo(); st.warning("Base de endotoxina limpa."); st.rerun()

    with col_ei:
        _cnae_labels = {
            "3250701": "Fab. equip. médico-odontológicos",
            "2660400": "Fab. equip. irradiação/imagem médica",
            "8129000": "Atividades de higienização/limpeza",
            "8650099": "Outros serviços de saúde humana",
            "2121101": "Farmácias de manipulação",
            "2122600": "Fabricação medicamentos humanos",
        }
        st.markdown(
            '<div style="background:#FFF;border:1.5px solid #E0E0E0;border-radius:14px;padding:1.2rem 1.4rem;">' +
            '<div style="font-weight:700;font-size:.9rem;color:#2D2D2D;margin-bottom:.8rem;">📋 CNAEs alvo e fontes</div>' +
            '<div style="display:flex;flex-direction:column;gap:.5rem;">',
            unsafe_allow_html=True
        )
        for idx, cnae_code in enumerate(cnaes_t5, 1):
            label = _cnae_labels.get(cnae_code, cnae_code)
            st.markdown(
                f'<div style="display:flex;gap:10px;align-items:center;">' +
                f'<span style="background:#EDE9FE;border-radius:50%;width:24px;height:24px;min-width:24px;' +
                f'display:flex;align-items:center;justify-content:center;font-size:.7rem;font-weight:700;color:#5B21B6;">{idx}</span>' +
                f'<div><strong style="font-size:.84rem;">{cnae_code}</strong>' +
                f'<span style="font-size:.78rem;color:#888;margin-left:6px;">{label}</span></div></div>',
                unsafe_allow_html=True
            )
        st.markdown(
            '</div><hr style="border:1px solid #F1F3F5;margin:.7rem 0;">' +
            '<div style="font-size:.8rem;color:#888;">' +
            '<strong>Ordem de prioridade:</strong><br>' +
            '1️⃣ <strong>Casa dos Dados</strong> — CNPJ oficial, razão social, telefone, email<br>' +
            '2️⃣ <strong>Econodata</strong> — fallback com scraping por CNAE/UF<br>' +
            '3️⃣ <strong>DuckDuckGo</strong> — apenas se ativado, somente empresas (bloqueia artigos)' +
            '</div></div>',
            unsafe_allow_html=True
        )

    if iniciar_endo:
        ufs_busca = (["SP","MG","SC","RS","PR","RJ","GO","CE","PE","DF","BA"]
                     if "Todos" in uf_endo else uf_endo)
        with st.status("🔬 Buscando empresas por CNAE...", expanded=True) as se:
            all_endo = []; seen_cnpj = set(); seen_nome = set()

            if "Casa dos Dados" in fonte_prim or "Econodata" in fonte_prim:
                prog_e = st.progress(0)
                total_steps = len(cnaes_t5) * len(ufs_busca[:4])
                step = 0
                for cnae_code in cnaes_t5:
                    for uf_item in ufs_busca[:4]:
                        step += 1
                        st.write(f"📊 CNAE {cnae_code} · {uf_item} ({'Casa dos Dados' if 'Casa' in fonte_prim else 'Econodata'})...")
                        try:
                            if "Casa dos Dados" in fonte_prim:
                                lotes = endo_bot.buscar_cnae(cnae_code, uf_item, max_pags)
                            else:
                                lotes = endo_bot.buscar_econodata(cnae_code, uf_item.lower())

                            for l in lotes:
                                # Dedup por CNPJ ou nome
                                key_cnpj = re.sub(r"\D","", l.get("cnpj",""))
                                key_nome = _slug(l.get("nome",""))
                                if (key_cnpj and key_cnpj in seen_cnpj) or key_nome in seen_nome:
                                    continue
                                if key_cnpj: seen_cnpj.add(key_cnpj)
                                if key_nome: seen_nome.add(key_nome)
                                l["fonte_busca"] = f"CNAE:{cnae_code}/{uf_item}"
                                l["escopo"]      = escopo_t5
                                # Cruzamento completo com Persona ICP
                                l = CRMAnalyzer.enrich_lead_from_crm(l, crm_df_all)
                                icp_s = CRMAnalyzer.icp_score_lead(l, crm_df_all)
                                l["crm_status"]  = icp_s.get("crm_status","novo")
                                l["lookalike"]   = icp_s.get("total", CRMAnalyzer.lookalike_score(l))
                                l["icp_acao"]    = icp_s.get("acao","prospectar")
                                l["icp_flags"]   = " | ".join(icp_s.get("flags",[]))
                                l["icp_prior"]   = icp_s.get("prioridade","normal")
                                db.upsert_endo(l); all_endo.append(l)

                            st.write(f"  ✅ {len(lotes)} empresas encontradas")
                        except Exception as e:
                            st.write(f"  ⚠️ {e}")
                        prog_e.progress(step / total_steps)
                        time.sleep(0.5)

            # DDG como complemento opcional
            if "DuckDuckGo" in fonte_prim and seeds_lista:
                st.write("🔍 DuckDuckGo (empresa-only)...")
                for seed in seeds_lista[:4]:
                    try:
                        lotes = endo_bot._ddg(seed, 8)
                        for l in lotes:
                            k = _slug(l.get("nome",""))
                            if k in seen_nome: continue
                            seen_nome.add(k)
                            l["fonte_busca"] = seed; l["escopo"] = escopo_t5
                            l = CRMAnalyzer.enrich_lead_from_crm(l, crm_df_all)
                            l["crm_status"]  = CRMAnalyzer.crm_status(
                                l.get("nome",""), crm_df_all, l.get("cnpj",""))
                            l["lookalike"]   = CRMAnalyzer.lookalike_score(l)
                            db.upsert_endo(l); all_endo.append(l)
                    except Exception as e:
                        st.write(f"  ⚠️ {e}")
                    time.sleep(1.2)

            se.update(label=f"✅ {len(all_endo)} empresas reais encontradas!", state="complete")
        st.success(f"**{len(all_endo)} empresas** salvas · CNAE {' / '.join(cnaes_t5)}"); st.rerun()


    st.markdown("<br>", unsafe_allow_html=True)
    endo_leads=db.get_endo_leads(); endo_st=db.get_endo_stats()
    me1,me2,me3 = st.columns(3,gap="small")
    with me1: st.markdown(f'<div class="mc"><div class="ml">Total</div><div class="mn mn-vl">{endo_st["total"]}</div></div>', unsafe_allow_html=True)
    with me2: st.markdown(f'<div class="mc"><div class="ml">Novos</div><div class="mn">{endo_st["novo"]}</div></div>', unsafe_allow_html=True)
    with me3: st.markdown(f'<div class="mc"><div class="ml">Contatados</div><div class="mn mn-am">{endo_st["contatado"]}</div></div>', unsafe_allow_html=True)

    if endo_leads:
        st.markdown("<br>", unsafe_allow_html=True)
        df_endo = pd.DataFrame(endo_leads)

        # ── Output estruturado: normalização e limpeza ────────────────
        def _clean(v):
            if v is None or str(v).strip() in ("","None","nan","NaN"): return "—"
            return str(v).strip()

        df_endo["nome"]      = df_endo["nome"].apply(_clean)
        df_endo["email"]     = df_endo["email"].apply(_clean)
        df_endo["telefone"]  = df_endo["telefone"].apply(_clean)
        df_endo["site"]      = df_endo["site"].apply(_clean)
        df_endo["crm_status"] = df_endo.get("crm_status", pd.Series(["—"]*len(df_endo))).fillna("—").apply(_clean)
        df_endo["lookalike"]  = pd.to_numeric(df_endo.get("lookalike", pd.Series([0]*len(df_endo))).fillna(0), errors="coerce").fillna(0).astype(int)
        df_endo["escopo"]     = df_endo.get("escopo", pd.Series(["—"]*len(df_endo))).fillna("—").apply(_clean)

        # CRM status badge
        CRM_ICON = {"vendida":"✅","perdida":"❌","em_andamento":"🔄","novo":"🆕","—":"🆕"}
        df_endo["CRM"] = df_endo["crm_status"].map(lambda x: CRM_ICON.get(x,"🆕")+" "+x.title().replace("_"," "))

        # Lookalike badge
        def _lla(score):
            if score>=70: return "🔥 Alto"
            if score>=40: return "⭐ Médio"
            if score>0:   return "➕ Baixo"
            return "—"
        df_endo["Lookalike"] = df_endo["lookalike"].apply(_lla)

        # Filtros inline
        fe1,fe2,fe3 = st.columns(3, gap="small")
        with fe1:
            f_crm = st.selectbox("Filtrar CRM", ["Todos","novo","vendida","perdida","em_andamento"], key="f_crm_endo")
        with fe2:
            f_esc = st.selectbox("Escopo", ["Todos"]+sorted(df_endo["escopo"].unique().tolist()), key="f_esc_endo")
        with fe3:
            f_lla = st.selectbox("Lookalike", ["Todos","🔥 Alto","⭐ Médio","➕ Baixo"], key="f_lla_endo")

        df_view = df_endo.copy()
        if f_crm  != "Todos": df_view = df_view[df_view["crm_status"]==f_crm]
        if f_esc  != "Todos": df_view = df_view[df_view["escopo"]==f_esc]
        if f_lla  != "Todos": df_view = df_view[df_view["Lookalike"]==f_lla]

        # Ordenar: Lookalike desc, crm_status (novo primeiro)
        crm_order = {"novo":0,"em_andamento":1,"perdida":2,"vendida":3,"—":4}
        df_view["_sort_crm"] = df_view["crm_status"].map(crm_order).fillna(4)
        df_view = df_view.sort_values(["lookalike","_sort_crm"], ascending=[False,True])

        cols_disp = ["nome","email","telefone","CRM","Lookalike","escopo","site","estado","cidade","status"]
        st.dataframe(df_view[[c for c in cols_disp if c in df_view.columns]].reset_index(drop=True),
                     use_container_width=True, height=380)

        st.markdown(f'<div style="font-size:.78rem;color:#888;margin:.3rem 0;">Exibindo <strong>{len(df_view)}</strong> de {len(df_endo)} leads · 🔥 Alto Lookalike = perfil semelhante aos clientes do CRM</div>', unsafe_allow_html=True)

        # ── Atualizar status ──────────────────────────────────────────
        st.markdown('<div style="font-weight:600;font-size:.88rem;color:#2D2D2D;margin:.6rem 0 .4rem;">⚙️ Atualizar status</div>', unsafe_allow_html=True)
        ce1,ce2,ce3 = st.columns([3,1,1],gap="small")
        with ce1:
            ne_sel = st.selectbox("Lead endo", df_view["nome"].tolist(), label_visibility="collapsed", key="sel_endo")
        with ce2:
            if st.button("✔️ Contatado", key="btn_econt", use_container_width=True):
                sel_e=next((l for l in endo_leads if l["nome"]==ne_sel),None)
                if sel_e: db.set_endo_status(sel_e["id"],"contatado"); st.rerun()
        with ce3:
            if st.button("✖️ Descartar", key="btn_edesc", use_container_width=True):
                sel_e=next((l for l in endo_leads if l["nome"]==ne_sel),None)
                if sel_e: db.set_endo_status(sel_e["id"],"descartado"); st.rerun()

        buf_e=io.BytesIO(); buf_e.write(exporter.generate_endo_excel())
        st.download_button("⬇  Exportar leads estruturados (.xlsx)", buf_e.getvalue(),
                           "leads_scitec_estruturado.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.markdown('<div class="empty"><div class="empty-i">🧪</div><div class="empty-t">Nenhum lead ainda.</div><div style="font-size:.78rem;color:#888;margin-top:4px;">Selecione um escopo e clique em <strong>🔬 Iniciar Busca</strong>.</div></div>', unsafe_allow_html=True)


# ═════════════════════════════════════════════════════════════════════
#  PERSONA ENGINE — Perfil ICP derivado do CRM + buscador inteligente
# =════════════════════════════════════════════════════════════════════

# CNAEs mapeados dos clientes-sucesso no CRM da Scitec
ICP_CNAES = {
    "3250701": "Fab. instrumentos e materiais médicos e odontológicos",
    "3250702": "Fab. artefatos médicos e odontológicos",
    "3250703": "Fab. equip. de raios X e radiodiagnóstico",
    "3250704": "Fab. materiais para medicina nuclear",
    "3250705": "Fab. materiais ópticos, fotográficos e cinematográficos",
    "3250706": "Fab. artefatos de higiene e/ou farmácia",
    "3250707": "Fab. outros equip. e instrumentos médicos",
    "3841000": "Fab. automóveis, camionetes e utilitários",  # não relevante
    "4645101": "Com. atacadista equip. médico-hospitalares",
    "4645102": "Com. atacadista próteses e artigos de ortopedia",
    "4773300": "Com. varejista artigos médicos e ortopédicos",
    "2660400": "Fab. instrumentos de medição e controle",
    "7120100": "Testes e análises técnicas",
    "2123800": "Fab. defensivos agrícolas",  # saneantes
    "2122600": "Fab. medicamentos para uso humano",
    "2121101": "Fab. analgésicos",
    "2121102": "Fab. medicamentos para uso veterinário",
}

# Perfil ICP fixo (derivado dos 32 negócios ganhos no CRM)
ICP_PROFILE = {
    "produtos":         ["ISO 10993 (Biocompatibilidade)", "MRI (Ressonância Magnética)"],
    "ticket_medio":     75245,
    "ticket_min":       11500,
    "ticket_max":       227016,
    "cargos_decisores": [
        "Gerente de Produção e Assuntos Regulatórios",
        "Engenharia de Materiais",
        "Qualidade",
        "Engenheiro Mecânico",
        "Analista de P&D",
        "Compras",
        "Diretor Técnico",
        "Gerente de P&D",
        "Responsável Regulatório",
    ],
    "segmentos": [
        "Fabricante implante ortopédico",
        "Fabricante implante dental/odontológico",
        "Fabricante dispositivo médico implantável",
        "Fabricante equipamento médico-hospitalar",
        "Fabricante material cirúrgico estéril",
        "Fabricante produto biomédico/biomaterial",
        "Fabricante prótese cardiovascular",
        "Importadora/distribuidora equip. médico",
    ],
    "cnaes_alvo": ["3250701","3250702","3250703","3250707","4645101","4645102"],
    "estados_alvo": ["SP","SC","MG","RS","PR","RJ"],
    "palavras_chave": [
        "implante","implant","ortopéd","ortoped","biomédic","biomed","odontológ",
        "dental","engenharia médic","equip médic","produto médic","biomaterial",
        "prótese","fixador","parafuso","placa ortopédica","endoprótese",
    ],
}


class PersonaEngine:
    """Motor de ICP: analisa CRM, gera personas e buscas inteligentes."""

    @staticmethod
    def build_icp(crm_df):
        """Constrói ICP dinâmico completo a partir dos dados do CRM."""
        if crm_df is None:
            return ICP_PROFILE.copy()

        icp = ICP_PROFILE.copy()

        # ── Segmentar por Estado ──────────────────────────────────────
        if "Estado" in crm_df.columns:
            vendidas = crm_df[crm_df["Estado"] == "Vendida"].copy()
            perdidas = crm_df[crm_df["Estado"] == "Perdida"].copy()
            andamento= crm_df[crm_df["Estado"] == "Em Andamento"].copy()
        else:
            vendidas = crm_df.copy()
            perdidas = pd.DataFrame()
            andamento= pd.DataFrame()

        if len(vendidas) == 0:
            vendidas = crm_df.copy()

        # ── Ticket médio dinâmico ─────────────────────────────────────
        if "Valor Único" in vendidas.columns:
            vals = pd.to_numeric(vendidas["Valor Único"], errors="coerce").dropna()
            if len(vals) > 0:
                icp["ticket_medio"] = int(round(vals.mean()))
                icp["ticket_min"]   = int(round(vals.min()))
                icp["ticket_max"]   = int(round(vals.max()))
                icp["ticket_total"] = int(round(vals.sum()))

        # ── Produtos (ISO 10993, MRI, etc.) ──────────────────────────
        if "Produtos" in vendidas.columns:
            icp["produtos_crm"] = vendidas["Produtos"].value_counts().to_dict()

        # ── Cargos decisores encontrados ──────────────────────────────
        if "Cargo" in vendidas.columns:
            cargos = [c.strip() for c in vendidas["Cargo"].tolist() if str(c).strip()]
            if cargos:
                icp["cargos_encontrados"] = sorted(set(cargos))

        # ── Contatos e emails dos clientes ganhos ─────────────────────
        if "Contatos" in vendidas.columns:
            contatos = [c.strip() for c in vendidas["Contatos"].tolist() if str(c).strip()]
            icp["contatos_crm"] = list(set(contatos))[:20]
        if "Email" in vendidas.columns:
            emails = [e.strip().lower() for e in vendidas["Email"].tolist()
                      if str(e).strip() and "@" in str(e)]
            icp["emails_crm"] = emails[:20]

        # ── Empresas de referência (vendidas) ─────────────────────────
        if "Empresa" in vendidas.columns:
            icp["empresas_ref"] = [e.strip() for e in vendidas["Empresa"].unique().tolist()
                                   if str(e).strip()][:15]

        # ── Fontes de origem dos leads ganhos ─────────────────────────
        if "Origem do Lead" in vendidas.columns:
            icp["origens_sucesso"] = vendidas["Origem do Lead"].value_counts().to_dict()

        # ── Taxas de conversão ────────────────────────────────────────
        total_neg = len(crm_df)
        total_v   = len(vendidas)
        total_p   = len(perdidas)
        if total_v + total_p > 0:
            icp["taxa_conversao"] = round(total_v / (total_v + total_p) * 100)
        icp["total_negocios"] = total_neg
        icp["total_vendidos"] = total_v
        icp["total_perdidos"] = total_p
        icp["total_andamento"]= len(andamento)

        # ── Empresas para reativação (perdidas com potencial) ─────────
        if len(perdidas) > 0 and "Motivo de Perda" in perdidas.columns:
            _NAO_REATIVAR = {"Demanda Nunca Existiu","Desistência","Preço",
                             "(Não Usar) Atendimento"}
            reativ = perdidas[~perdidas["Motivo de Perda"].isin(_NAO_REATIVAR)]
            cols_reativ = [c for c in ["Empresa","Produtos","Valor Único","Motivo de Perda"]
                           if c in reativ.columns]
            icp["reativacao"] = reativ[cols_reativ].to_dict("records")[:15]

        # ── Motivos de perda para análise ─────────────────────────────
        if len(perdidas) > 0 and "Motivo de Perda" in perdidas.columns:
            icp["motivos_perda"] = perdidas["Motivo de Perda"].value_counts().to_dict()

        return icp

    @staticmethod
    def generate_smart_seeds(icp, produto_foco=None, uf=None):
        """Gera seeds otimizadas com base no ICP."""
        seeds = []
        uf_suffix = f" {uf}" if uf else " Brasil"

        if produto_foco == "ISO10993" or produto_foco is None:
            for seg in icp.get("segmentos", [])[:4]:
                seeds.append(f"{seg} ISO 10993 ANVISA{uf_suffix}")
            seeds += [
                f"fabricante implante ortopédico biocompatibilidade registro ANVISA{uf_suffix}",
                f"empresa dispositivo médico implantável ensaio biológico{uf_suffix}",
                f"indústria biomédica ortopedia implante titânio ANVISA{uf_suffix}",
            ]

        if produto_foco == "MRI" or produto_foco is None:
            seeds += [
                f"fabricante implante compatível ressonância magnética MRI{uf_suffix}",
                f"fabricante dispositivo médico MRI safe implante ortopédico{uf_suffix}",
                f"indústria implante MRI conditional artefato{uf_suffix}",
            ]

        if produto_foco == "BIOBURDEN" or produto_foco is None:
            seeds += [
                f"farmácia manipulação injetáveis ANVISA registro{uf_suffix}",
                f"indústria farmacêutica fabricante produto injetável parenterais{uf_suffix}",
                f"fabricante dispositivo médico estéril bioburden ANVISA{uf_suffix}",
            ]

        # Seeds de lookalike baseadas em empresas de referência
        for kw in icp.get("palavras_chave", [])[:4]:
            seeds.append(f"empresa {kw} fabricante registro ANVISA{uf_suffix}")

        return list(dict.fromkeys(seeds))  # deduplica mantendo ordem

    @staticmethod
    def econodata_search_urls(cnaes, ufs=None):
        """Gera URLs do Econodata por CNAE e UF."""
        urls = []
        ufs_list = ufs or ICP_PROFILE["estados_alvo"]
        for cnae in cnaes:
            for uf in ufs_list[:3]:
                urls.append({
                    "url": f"https://www.econodata.com.br/empresas/{uf.lower()}/{cnae}/1",
                    "cnae": cnae,
                    "uf": uf,
                    "desc": ICP_CNAES.get(cnae, cnae),
                })
        return urls


class ContactHunter:
    """Enriquecimento multi-fonte: Casa dos Dados, Econodata, LinkedIn/DDG."""

    def __init__(self):
        self._s = requests.Session()
        self._s.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "pt-BR,pt;q=0.9,en;q=0.8",
            "Accept-Encoding": "gzip, deflate, br",
        })
        self._re_email = re.compile(
            r"[a-zA-Z0-9][a-zA-Z0-9._%+\-]{1,40}@[a-zA-Z0-9.\-]{2,40}\.[a-zA-Z]{2,6}"
        )
        self._re_fone = re.compile(
            r"(?:\(?0?[1-9]{2}\)?\s?|0?[1-9]{2}[\s\-])"
            r"(?:9[\s]?[1-9]\d{3}|[2-8]\d{3})[\s\-]?\d{4}"
        )

    # ── 1. Casa dos Dados ─────────────────────────────────────────────
    def casa_dados_search(self, razao_social="", cnae="", uf="", page=1):
        """Busca empresas na API pública Casa dos Dados."""
        try:
            payload = {"query": {}, "page": page}
            if razao_social: payload["query"]["razao_social"] = razao_social
            if cnae:         payload["query"]["cnae_fiscal"]  = int(cnae) if cnae.isdigit() else cnae
            if uf:           payload["query"]["uf"]            = uf.upper()
            r = self._s.post(
                "https://api.casadosdados.com.br/v2/public/cnpj/search",
                json=payload, timeout=15
            )
            if r.status_code == 200:
                data = r.json()
                return [self._norm_casadadados(item) for item in data.get("data", [])]
        except Exception as e:
            logger.warning(f"Casa dos Dados: {e}")
        return []

    def _norm_casadadados(self, item):
        cnpj_raw = re.sub(r"\D","",item.get("cnpj",""))
        c = cnpj_raw
        cnpj_fmt = f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}" if len(cnpj_raw)==14 else cnpj_raw
        return {
            "cnpj":     cnpj_fmt,
            "nome":     item.get("razao_social","") or item.get("nome_fantasia",""),
            "email":    (item.get("email","") or "").lower() or None,
            "telefone": self._fmt_fone(item.get("telefone1","") or item.get("ddd1","")+item.get("telefone1","")),
            "cidade":   item.get("municipio",""),
            "estado":   item.get("uf",""),
            "setor":    item.get("cnae_fiscal_descricao","") or item.get("atividade_principal",""),
            "cnae":     str(item.get("cnae_fiscal","") or ""),
            "fonte":    "casa_dados",
        }

    # ── 2. Econodata (CNAE search) ────────────────────────────────────
    def econodata_cnae(self, cnae, uf="sp", page=1):
        """Scraping do Econodata por CNAE e UF — com validação _is_empresa."""
        cnae_clean = re.sub(r"\D","", str(cnae))
        try:
            url = f"https://www.econodata.com.br/empresas/{uf.lower()}/{cnae_clean}/{page}"
            r   = self._s.get(url, timeout=15)
            if r.status_code != 200: return []
            soup  = BeautifulSoup(r.text, "html.parser")
            leads = []

            # Tenta múltiplos seletores em ordem de especificidade
            _SELECTORS = [
                ("a.company-result", ".company-name", ".cnpj", ".city"),
                (".empresa-item",    "h3",             ".cnpj", ".municipio"),
                (".company-card",    ".company-name",  ".cnpj", ".cidade"),
                (".listing-item",    "h2",             None,    None),
                ("article.empresa",  "h2",             ".cnpj", ".cidade"),
            ]
            for container_sel, nome_sel, cnpj_sel, cidade_sel in _SELECTORS:
                cards = soup.select(container_sel)
                if not cards: continue
                for card in cards[:30]:
                    ne = card.select_one(nome_sel) if nome_sel else card
                    if not ne: continue
                    nome = ne.get_text(strip=True)
                    if not nome or not _is_empresa(nome): continue
                    cnpj_raw = ""
                    if cnpj_sel:
                        ce = card.select_one(cnpj_sel)
                        if ce: cnpj_raw = re.sub(r"\D","", ce.get_text())
                    # Só aceita se CNPJ tem 14 dígitos OU nome é claramente empresa
                    if len(cnpj_raw) == 14 or _is_empresa(nome):
                        c = cnpj_raw
                        cnpj_fmt = (f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}"
                                    if len(c)==14 else c)
                        cidade = ""
                        if cidade_sel:
                            cde = card.select_one(cidade_sel)
                            if cde: cidade = cde.get_text(strip=True)
                        leads.append({
                            "cnpj": cnpj_fmt, "nome": nome, "cidade": cidade,
                            "estado": uf.upper(), "cnae": cnae_clean,
                            "site": None, "snippet": f"CNAE {cnae_clean} — {uf.upper()}",
                            "fonte": "econodata", "email": None,
                            "telefone": None, "decisor": None,
                        })
                if leads: break

            # Fallback: JSON-LD schema.org embutido na página
            if not leads:
                import json
                for script in soup.find_all("script", {"type":"application/ld+json"}):
                    try:
                        data = json.loads(script.string or "")
                        items = data if isinstance(data,list) else [data]
                        for item in items:
                            if item.get("@type") in ("LocalBusiness","Organization","Corporation"):
                                nome = item.get("name","").strip()
                                if nome and _is_empresa(nome):
                                    leads.append({
                                        "cnpj": "", "nome": nome,
                                        "email": item.get("email","") or None,
                                        "telefone": item.get("telephone","") or None,
                                        "cidade": "",  "estado": uf.upper(),
                                        "cnae": cnae_clean, "site": item.get("url","") or None,
                                        "snippet": item.get("description","")[:100],
                                        "fonte": "econodata_jsonld", "decisor": None,
                                    })
                    except Exception:
                        pass

            return leads
        except Exception as e:
            logger.warning(f"Econodata {cnae}/{uf}: {e}")
            return []

    # ── 3. Conodata ───────────────────────────────────────────────────
    def conodata_cnpj(self, cnpj):
        """Busca dados de contato no Conodata pelo CNPJ."""
        cnpj_clean = re.sub(r"\D","",cnpj)
        if len(cnpj_clean)!=14: return {}
        try:
            r = self._s.get(f"https://conodata.com.br/empresa/cnpj/{cnpj_clean}", timeout=12)
            if r.status_code != 200: return {}
            soup = BeautifulSoup(r.text, "html.parser")
            result = {}
            # Email
            emails = [e.lower() for e in self._re_email.findall(r.text)
                      if not any(bl in e for bl in ("sentry","github","@gov","@serasaexp","png","jpg"))]
            if emails: result["email"] = emails[0]
            # Telefone
            fones = self._re_fone.findall(r.text)
            if fones: result["telefone"] = self._fmt_fone(fones[0])
            # Decisor
            for el in soup.select(".contato, .responsavel, .decisor, [class*=contact], [class*=pessoa]"):
                txt = el.get_text(" ", strip=True)
                if re.search(r"[A-ZÁÉÍÓÚ][a-záéíóú]+ [A-ZÁÉÍÓÚ][a-záéíóú]+", txt):
                    result["decisor"] = txt[:80]
                    break
            return result
        except Exception as e:
            logger.warning(f"Conodata {cnpj}: {e}")
            return {}

    # ── 4. LinkedIn via DuckDuckGo ────────────────────────────────────
    def linkedin_decisor(self, empresa_nome, cargo_hint="gerente OR diretor OR qualidade"):
        """Busca decisor no LinkedIn via DuckDuckGo HTML."""
        try:
            query = f'site:linkedin.com/in "{empresa_nome}" ({cargo_hint})'
            r = self._s.post("https://html.duckduckgo.com/html/",
                             data={"q": query, "kl": "br-pt"}, timeout=15)
            if r.status_code != 200: return None
            soup = BeautifulSoup(r.text, "html.parser")
            for res in soup.select(".result")[:3]:
                title = (res.select_one(".result__title") or {}).get_text(strip=True) if res.select_one(".result__title") else ""
                snip  = (res.select_one(".result__snippet") or {}).get_text(strip=True) if res.select_one(".result__snippet") else ""
                # Extrai nome da pessoa do título LinkedIn (formato: "Nome Sobrenome - Cargo - Empresa")
                m = re.match(r"^([A-ZÁÉÍÓÚ][a-záéíóú]+(?:\s[A-ZÁÉÍÓÚ][a-záéíóú]+){1,3})", title, re.UNICODE)
                if m:
                    nome_pessoa = m.group(1)
                    cargo_match = re.search(r"-\s*([^-|]{5,50})\s*[-|]", title)
                    cargo_txt   = cargo_match.group(1).strip() if cargo_match else ""
                    return {"decisor": nome_pessoa, "cargo": cargo_txt, "fonte_decisor": "linkedin"}
        except Exception as e:
            logger.warning(f"LinkedIn DDG: {e}")
        return None

    # ── 5. Pipeline completo para um lead ────────────────────────────
    def enrich_pipeline(self, lead, crm_df=None, status_cb=None):
        """Aplica todas as fontes em cascata e atualiza o lead."""
        nome  = lead.get("nome","")
        cnpj  = lead.get("cnpj","")
        etapa = []

        if status_cb: status_cb(f"🔍 Enriquecendo **{nome[:40]}**...")

        # CRM check
        if crm_df is not None:
            crm_st = CRMAnalyzer.crm_status(nome, crm_df)
            lead["crm_status"] = crm_st
            detail  = CRMAnalyzer.crm_detail(nome, crm_df)
            if detail:
                lead.update({k:v for k,v in detail.items() if v and not lead.get(k)})
            etapa.append(f"CRM:{crm_st}")

        # Lookalike
        lead["lookalike"] = CRMAnalyzer.lookalike_score(lead)

        # Conodata (contatos)
        if cnpj and (not lead.get("email") or not lead.get("telefone")):
            cono = self.conodata_cnpj(cnpj)
            if cono.get("email")    and not lead.get("email"):    lead["email"]    = cono["email"]
            if cono.get("telefone") and not lead.get("telefone"): lead["telefone"] = cono["telefone"]
            if cono.get("decisor")  and not lead.get("decisor"):  lead["decisor"]  = cono["decisor"]
            if cono: etapa.append("Conodata:OK")

        # LinkedIn decisor
        if not lead.get("decisor") and nome:
            cargo_hint = " OR ".join(f'"{c}"' for c in ICP_PROFILE["cargos_decisores"][:5])
            li = self.linkedin_decisor(nome, cargo_hint)
            if li:
                lead["decisor"] = li.get("decisor","")
                if li.get("cargo"): lead.setdefault("cargo_decisor", li["cargo"])
                etapa.append("LinkedIn:OK")

        lead["_fontes"] = " · ".join(etapa)
        return lead

    def _fmt_fone(self, raw):
        if not raw: return None
        d = re.sub(r"\D","",str(raw))
        if len(d)==11: return f"({d[:2]}) {d[2]}.{d[3:7]}-{d[7:]}"
        if len(d)==10: return f"({d[:2]}) {d[2:6]}-{d[6:]}"
        return raw if raw.strip() else None


# Instâncias globais
persona_engine  = PersonaEngine()
contact_hunter  = ContactHunter()


# ══════════════════════════════════════════════════════════════════════
#  ABA 6 — PERSONA ICP (Ideal Customer Profile)
# ══════════════════════════════════════════════════════════════════════
with t6:
    st.markdown(
        f'<div class="sec-title">🎯 Persona ICP — {"Laboratório" if IS_LAB else "Certificadora OCP"}</div>',
        unsafe_allow_html=True
    )
    _mod_desc = ("Perfil ICP baseado em ensaios biológicos · Bioburden · Endotoxina · ISO 10993 · MRI"
                 if IS_LAB else
                 "Perfil ICP baseado em certificações OCP · Portaria 145 · Portaria 384 · INMETRO")
    st.markdown(f'<div class="sec-desc">{_mod_desc} · Cruzamento CRM automático.</div>',
                unsafe_allow_html=True)

    # ── Construir ICP do CRM ─────────────────────────────────────────
    icp = persona_engine.build_icp(crm_df_all)

    # ════ SEÇÃO 1: PERFIL ICP ════════════════════════════════════════
    with st.expander("📊 Perfil do Cliente Ideal (derivado do CRM)", expanded=True):
        p1,p2,p3 = st.columns(3, gap="medium")

        with p1:
            st.markdown('<div style="font-weight:700;font-size:.88rem;color:#2D2D2D;margin-bottom:.6rem;">💰 Tickets & Conversão</div>', unsafe_allow_html=True)
            t_med  = icp.get("ticket_medio",75245)
            t_min  = icp.get("ticket_min",11500)
            t_max  = icp.get("ticket_max",227016)
            t_tot  = icp.get("ticket_total",0)
            taxa   = icp.get("taxa_conversao",0)
            tv     = icp.get("total_vendidos",0)
            tp     = icp.get("total_perdidos",0)
            ta     = icp.get("total_andamento",0)
            st.markdown(
                f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-bottom:.6rem;">' +
                f'<div class="mc" style="min-height:65px;"><div class="ml">Ticket Médio</div><div class="mn" style="font-size:1.3rem;">R$ {t_med:,.0f}</div></div>' +
                f'<div class="mc" style="min-height:65px;"><div class="ml">Conversão</div><div class="mn mn-am" style="font-size:1.3rem;">{taxa}%</div></div>' +
                f'<div class="mc" style="min-height:65px;"><div class="ml">Ganhos ✅</div><div class="mn" style="font-size:1.3rem;color:#16a34a;">{tv}</div></div>' +
                f'<div class="mc" style="min-height:65px;"><div class="ml">Perdidos ❌</div><div class="mn mn-hot" style="font-size:1.3rem;">{tp}</div></div>' +
                '</div>',
                unsafe_allow_html=True
            )
            st.markdown(f'<div style="font-size:.77rem;color:#888;">Min: R$ {t_min:,.0f} · Max: R$ {t_max:,.0f} · Total vendido: R$ {t_tot:,.0f}</div>', unsafe_allow_html=True)
            if "produtos_crm" in icp:
                st.markdown('<div style="font-weight:600;font-size:.82rem;color:#2D2D2D;margin:.7rem 0 .3rem;">📦 Produtos no CRM</div>', unsafe_allow_html=True)
                total_prods = sum(icp["produtos_crm"].values())
                for prod, n in icp["produtos_crm"].items():
                    pct = round(n/total_prods*100)
                    cor = "#7C3AED" if "10993" in str(prod) else "#0D9291" if "MRI" in str(prod) else "#3B82F6"
                    st.markdown(f'<div style="display:flex;align-items:center;gap:6px;margin-bottom:4px;font-size:.82rem;"><span style="min-width:120px;font-weight:600;">{prod[:15]}</span><div class="progress-bar-wrap" style="flex:1;"><div class="progress-bar" style="width:{pct}%;background:{cor};"></div></div><span style="color:#888;min-width:35px;text-align:right;">{n}x</span></div>', unsafe_allow_html=True)
            if "motivos_perda" in icp:
                st.markdown('<div style="font-weight:600;font-size:.82rem;color:#2D2D2D;margin:.7rem 0 .3rem;">⚠️ Motivos de Perda</div>', unsafe_allow_html=True)
                for motivo, n in list(icp["motivos_perda"].items())[:5]:
                    st.markdown(f'<div style="font-size:.78rem;padding:3px 0;border-bottom:1px solid #F1F3F5;display:flex;justify-content:space-between;"><span>{motivo[:30]}</span><span style="color:#E53935;font-weight:600;">{n}x</span></div>', unsafe_allow_html=True)

        with p2:
            st.markdown('<div style="font-weight:700;font-size:.88rem;color:#2D2D2D;margin-bottom:.6rem;">🎯 Cargos dos Decisores</div>', unsafe_allow_html=True)
            cargos_show = icp.get("cargos_encontrados") or icp.get("cargos_decisores",[])
            for cargo in cargos_show[:8]:
                st.markdown(f'<div style="background:#F1F3F5;border-radius:6px;padding:5px 10px;font-size:.8rem;margin-bottom:4px;border-left:3px solid #0D9291;">👤 {cargo}</div>', unsafe_allow_html=True)
            # Contatos reais encontrados no CRM
            if icp.get("contatos_crm"):
                st.markdown('<div style="font-weight:600;font-size:.82rem;color:#2D2D2D;margin:.7rem 0 .3rem;">📋 Contatos reais (CRM)</div>', unsafe_allow_html=True)
                for contato in icp["contatos_crm"][:6]:
                    st.markdown(f'<div style="background:#E0F4F4;border-radius:6px;padding:4px 8px;font-size:.78rem;margin-bottom:3px;">👤 {contato}</div>', unsafe_allow_html=True)
            # Origens dos leads ganhos
            if icp.get("origens_sucesso"):
                st.markdown('<div style="font-weight:600;font-size:.82rem;color:#2D2D2D;margin:.7rem 0 .3rem;">📡 Origem dos Leads Ganhos</div>', unsafe_allow_html=True)
                for orig, n in list(icp["origens_sucesso"].items())[:5]:
                    if orig.strip():
                        st.markdown(f'<div style="font-size:.78rem;padding:3px 0;border-bottom:1px solid #F1F3F5;display:flex;justify-content:space-between;"><span>{orig[:35]}</span><span style="color:#0D9291;font-weight:600;">{n}x</span></div>', unsafe_allow_html=True)

        with p3:
            st.markdown('<div style="font-weight:700;font-size:.88rem;color:#2D2D2D;margin-bottom:.6rem;">🏭 Segmentos-alvo</div>', unsafe_allow_html=True)
            for seg in icp.get("segmentos",[])[:5]:
                st.markdown(f'<div style="background:#F1F3F5;border-radius:6px;padding:5px 10px;font-size:.8rem;margin-bottom:4px;border-left:3px solid #7C3AED;">🏭 {seg}</div>', unsafe_allow_html=True)
            st.markdown('<div style="font-weight:700;font-size:.82rem;color:#2D2D2D;margin:.6rem 0 .3rem;">📍 Estados prioritários</div>', unsafe_allow_html=True)
            st.markdown(', '.join(f'`{uf}`' for uf in icp.get("estados_alvo",[])))
            # Emails disponíveis no CRM
            if icp.get("emails_crm"):
                st.markdown('<div style="font-weight:600;font-size:.82rem;color:#2D2D2D;margin:.7rem 0 .3rem;">✉️ Emails CRM disponíveis</div>', unsafe_allow_html=True)
                for email in icp["emails_crm"][:5]:
                    if email and "@" in email:
                        dom = email.split("@")[1] if "@" in email else ""
                        st.markdown(f'<div style="font-size:.76rem;padding:3px 6px;background:#F8F9FA;border-radius:4px;margin-bottom:3px;font-family:monospace;">✉ {email[:40]}</div>', unsafe_allow_html=True)

        if icp.get("empresas_ref"):
            st.markdown('<div style="font-weight:600;font-size:.82rem;color:#2D2D2D;margin-top:.8rem;margin-bottom:.4rem;">✅ Clientes ganhos (referência ICP)</div>', unsafe_allow_html=True)
            # Enriquecer com status e produto do CRM
            cols_ref = st.columns(3, gap="small")
            for i, emp in enumerate(icp["empresas_ref"][:12]):
                with cols_ref[i%3]:
                    crm_d = CRMAnalyzer.crm_detail(emp, crm_df_all)
                    prod  = crm_d.get("crm_produto","") or ""
                    cor_prod = "#7C3AED" if "10993" in prod else "#0D9291" if "MRI" in prod else "#3B82F6"
                    badge = f'<span style="background:{cor_prod};color:#fff;font-size:.68rem;padding:1px 5px;border-radius:3px;margin-left:3px;">{prod[:10]}</span>' if prod else ""
                    st.markdown(
                        f'<div style="background:#E0F4F4;border-radius:6px;padding:4px 8px;font-size:.76rem;margin-bottom:3px;" title="{emp}">' +
                        f'<span style="overflow:hidden;text-overflow:ellipsis;white-space:nowrap;">{emp[:28]}</span>{badge}</div>',
                        unsafe_allow_html=True
                    )

    # ════ SEÇÃO 2: EMPRESAS PARA REATIVAR ════════════════════════════
    if icp.get("reativacao"):
        with st.expander(f"♻️ {len(icp['reativacao'])} Empresas para Reativação (perdidas com potencial)", expanded=False):
            df_reativ = pd.DataFrame(icp["reativacao"])
            df_reativ.columns = [c.replace("crm_","") for c in df_reativ.columns]
            df_reativ["Valor Único"] = pd.to_numeric(df_reativ.get("Valor Único",pd.Series()), errors="coerce").fillna(0).apply(lambda x: f"R$ {x:,.0f}" if x>0 else "—")
            st.dataframe(df_reativ, use_container_width=True, height=300)
            st.caption("💡 Foco: motivos de perda por Capacidade de Atendimento e Demanda Futura — podem ser abordados novamente.")

    st.markdown("<hr style='border:1px solid #E0E0E0;margin:1rem 0;'>", unsafe_allow_html=True)

    # ════ SEÇÃO 3: BUSCA INTELIGENTE POR CNAE ════════════════════════
    st.markdown('<div style="font-weight:700;font-size:.95rem;color:#2D2D2D;margin-bottom:.5rem;">🔍 Busca Inteligente por CNAE</div>', unsafe_allow_html=True)
    st.markdown('<div style="font-size:.82rem;color:#888;margin-bottom:.8rem;">Localiza fabricantes de dispositivos médicos diretamente no banco de dados CNPJ por código de atividade econômica.</div>', unsafe_allow_html=True)

    bc1,bc2,bc3,bc4 = st.columns([2,1,1,1], gap="small")
    with bc1:
        cnae_opcoes = {f"{cod} — {desc[:40]}": cod for cod,desc in ICP_CNAES.items()
                       if cod in ICP_PROFILE["cnaes_alvo"]}
        cnae_opcoes["Outro (digitar)"] = "custom"
        cnae_sel = st.selectbox("CNAE", list(cnae_opcoes.keys()), key="cnae_sel")
        cnae_cod = cnae_opcoes[cnae_sel]
        if cnae_cod == "custom":
            cnae_cod = st.text_input("Código CNAE", placeholder="3250701", key="cnae_custom")
    with bc2:
        uf_busca = st.selectbox("Estado", ["SP","MG","SC","RS","PR","RJ","GO","CE","PE","DF","Todos"], key="uf_busca_cnae")
    with bc3:
        fonte_busca_cnae = st.selectbox("Fonte", ["Casa dos Dados","Econodata","Ambas"], key="fonte_cnae")
    with bc4:
        st.markdown("<div style='margin-top:1.6rem;'>", unsafe_allow_html=True)
        buscar_cnae = st.button("🔎 Buscar por CNAE", key="btn_cnae_search", use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

    if buscar_cnae and cnae_cod and cnae_cod != "custom":
        with st.status("🔍 Buscando por CNAE...", expanded=True) as sc_cnae:
            cnae_results = []
            ufs_to_search = ICP_PROFILE["estados_alvo"] if uf_busca=="Todos" else [uf_busca]

            if fonte_busca_cnae in ("Casa dos Dados","Ambas"):
                st.write(f"📊 Casa dos Dados — CNAE {cnae_cod}...")
                for uf_item in ufs_to_search[:4]:
                    batch = contact_hunter.casa_dados_search(cnae=cnae_cod, uf=uf_item)
                    cnae_results.extend(batch)
                    if batch: st.write(f"  ✅ {uf_item}: {len(batch)} empresas")
                    time.sleep(0.4)

            if fonte_busca_cnae in ("Econodata","Ambas"):
                st.write(f"📊 Econodata — CNAE {cnae_cod}...")
                for uf_item in ufs_to_search[:3]:
                    batch = contact_hunter.econodata_cnae(cnae_cod, uf_item.lower())
                    cnae_results.extend(batch)
                    if batch: st.write(f"  ✅ {uf_item}: {len(batch)} empresas")
                    time.sleep(0.6)

            # Deduplica por CNPJ + nome, enriquece com CRM
            seen_cnpj_p = set(); seen_nome_p = set()
            cnae_unique = []
            for lead in cnae_results:
                k_cnpj = re.sub(r"\D","", lead.get("cnpj",""))
                k_nome = _slug(lead.get("nome",""))
                if (k_cnpj and k_cnpj in seen_cnpj_p) or (k_nome and k_nome in seen_nome_p):
                    continue
                if k_cnpj: seen_cnpj_p.add(k_cnpj)
                if k_nome: seen_nome_p.add(k_nome)
                # Cruzar com CRM: email, telefone, decisor quando disponível
                lead = CRMAnalyzer.enrich_lead_from_crm(lead, crm_df_all)
                icp_r = CRMAnalyzer.icp_score_lead(lead, crm_df_all)
                lead["crm_status"]  = icp_r.get("crm_status","novo")
                lead["lookalike"]   = icp_r.get("total", CRMAnalyzer.lookalike_score(lead))
                lead["icp_acao"]    = icp_r.get("acao","prospectar")
                lead["icp_flags"]   = " | ".join(icp_r.get("flags",[]))
                lead["icp_prior"]   = icp_r.get("prioridade","normal")
                cnae_unique.append(lead)
            # Ordena: novos primeiro, maior lookalike no topo
            cnae_unique.sort(key=lambda x: (
                {"novo":0,"em_andamento":1,"perdida":2,"vendida":3}.get(x.get("crm_status","novo"),0),
                -(x.get("lookalike",0) or 0)
            ))
            sc_cnae.update(label=f"✅ {len(cnae_unique)} empresas únicas encontradas!", state="complete")
            st.session_state["persona_cnae_results"] = cnae_unique


    # Exibir resultados CNAE
    cnae_res = st.session_state.get("persona_cnae_results", [])
    if cnae_res:
        df_cnae = pd.DataFrame(cnae_res)
        CRM_ICON = {"vendida":"✅ Vendida","perdida":"❌ Perdida",
                    "em_andamento":"🔄 Em Andamento","novo":"🆕 Novo","—":"🆕 Novo"}
        df_cnae["CRM"]       = df_cnae.get("crm_status", pd.Series(["novo"]*len(df_cnae))).fillna("novo").map(CRM_ICON)
        df_cnae["Lookalike"] = pd.to_numeric(df_cnae.get("lookalike",pd.Series([0]*len(df_cnae))),errors="coerce").fillna(0).astype(int).apply(lambda s: "🔥 Alto" if s>=70 else ("⭐ Médio" if s>=40 else ("➕ Baixo" if s>0 else "—")))
        # Mostrar decisor e produto do CRM quando disponível
        df_cnae["Decisor"]   = df_cnae.get("decisor",   pd.Series([""]*len(df_cnae))).fillna("").apply(lambda x: x if x and x!="None" else "—")
        df_cnae["Produto CRM"]= df_cnae.get("crm_produto",pd.Series([""]*len(df_cnae))).fillna("").apply(lambda x: x if x else "—")
        # Coluna Ação ICP
        _ACAO_ICON = {"upsell":"🚀 Upsell","reativar":"♻️ Reativar","nurturing":"🔄 Nurturing",
                      "prospectar":"🆕 Prospectar","descartar":"🗑 Descartar"}
        df_cnae["Ação ICP"]  = df_cnae.get("icp_acao", pd.Series(["prospectar"]*len(df_cnae))).fillna("prospectar").map(_ACAO_ICON)
        df_cnae["Flags ICP"] = df_cnae.get("icp_flags", pd.Series([""]*len(df_cnae))).fillna("")
        # Ordenação inteligente: prioridade alta primeiro, maior ICP no topo
        _PRIOR_ORDER = {"alta":0,"media":1,"normal":2}
        _sort_crm_p = {"🆕 Novo":0,"🔄 Em Andamento":1,"❌ Perdida":2,"✅ Vendida":3}
        df_cnae["_sort_prior"] = df_cnae.get("icp_prior",pd.Series(["normal"]*len(df_cnae))).fillna("normal").map(_PRIOR_ORDER).fillna(2)
        df_cnae["_sort"]       = df_cnae["CRM"].map(_sort_crm_p).fillna(0)
        df_cnae_sorted = df_cnae.sort_values(["_sort_prior","_sort","lookalike"],
                                              ascending=[True, True, False])
        cols_cn = ["nome","cnpj","email","telefone","Decisor","Ação ICP","Produto CRM","cidade","estado","CRM","Lookalike"]
        n_com_email    = int((df_cnae_sorted["email"].apply(lambda x: bool(x and x!="—" and "@" in str(x)))).sum())
        n_com_decisor  = int((df_cnae_sorted["Decisor"] != "—").sum())
        n_novos        = int((df_cnae_sorted["CRM"] == "🆕 Novo").sum())
        st.markdown(
            f'<div style="display:flex;gap:10px;margin-bottom:.5rem;flex-wrap:wrap;">' +
            f'<span style="background:#E0F4F4;color:#0A7170;border-radius:6px;padding:3px 9px;font-size:.78rem;font-weight:600;">🆕 {n_novos} novos</span>' +
            f'<span style="background:#EDE9FE;color:#5B21B6;border-radius:6px;padding:3px 9px;font-size:.78rem;font-weight:600;">✉️ {n_com_email} com e-mail</span>' +
            f'<span style="background:#FFF8E1;color:#92400E;border-radius:6px;padding:3px 9px;font-size:.78rem;font-weight:600;">👤 {n_com_decisor} com decisor</span>' +
            f'<span style="font-size:.78rem;color:#888;padding:3px 0;">Total: {len(cnae_res)} empresas</span>' +
            '</div>',
            unsafe_allow_html=True
        )
        st.dataframe(
            df_cnae_sorted[[c for c in cols_cn if c in df_cnae_sorted.columns]].drop(columns=["_sort"],errors="ignore").reset_index(drop=True),
            use_container_width=True, height=380
        )
        # Botão salvar na base
        if st.button("💾 Salvar todos na base Endotoxina", key="btn_save_cnae", use_container_width=False):
            saved = 0
            for lead in cnae_res:
                lead["escopo"]     = cnae_cod
                lead["fonte_busca"]= f"CNAE:{cnae_cod}"
                lead["snippet"]    = lead.get("setor","")
                db.upsert_endo(lead); saved += 1
            st.success(f"✅ {saved} empresas salvas na base Endotoxina."); st.rerun()

        buf_cn = io.BytesIO(); df_cnae_sorted.to_excel(buf_cn, index=False)
        st.download_button("⬇  Exportar resultados CNAE (.xlsx)", buf_cn.getvalue(),
                           f"cnae_{cnae_cod}_scitec.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    st.markdown("<hr style='border:1px solid #E0E0E0;margin:1rem 0;'>", unsafe_allow_html=True)

    # ════ SEÇÃO 4: BUSCA SEMÂNTICA COM SEEDS GERADAS PELO ICP ════════
    st.markdown('<div style="font-weight:700;font-size:.95rem;color:#2D2D2D;margin-bottom:.5rem;">🧠 Busca Semântica — Seeds Geradas pelo ICP</div>', unsafe_allow_html=True)
    st.markdown('<div style="font-size:.82rem;color:#888;margin-bottom:.8rem;">Seeds otimizadas para o perfil dos seus clientes reais. Busca via DuckDuckGo com enriquecimento multi-fonte.</div>', unsafe_allow_html=True)

    bs1,bs2,bs3 = st.columns([1,1,2], gap="small")
    with bs1:
        prod_foco = st.selectbox("Produto foco", ["Todos","ISO10993","MRI","BIOBURDEN"], key="prod_foco_persona")
    with bs2:
        uf_sem  = st.selectbox("Estado", ["Todos","SP","MG","SC","RS","PR","RJ","GO","CE"], key="uf_sem")
    with bs3:
        enrich_li = st.checkbox("🔍 Buscar decisor no LinkedIn (via DuckDuckGo)", value=True, key="enrich_li")

    pf = None if prod_foco=="Todos" else prod_foco
    uf_s = None if uf_sem=="Todos" else uf_sem
    smart_seeds = persona_engine.generate_smart_seeds(icp, pf, uf_s)

    st.text_area("Seeds geradas pelo ICP (editável):", value="\n".join(smart_seeds), height=130, key="persona_seeds_area")
    seeds_pers = [s.strip() for s in st.session_state.get("persona_seeds_area","").splitlines() if s.strip()]
    max_s = st.slider("Resultados por seed", 3, 12, 5, key="max_s_pers")
    buscar_sem = st.button("🚀 Iniciar Busca Semântica ICP", key="btn_sem_icp", use_container_width=False)

    if buscar_sem:
        with st.status("🔍 Busca semântica ICP em andamento...", expanded=True) as ss:
            sem_results = []; seen_sem = set()
            prog_s = st.progress(0)
            for i, seed in enumerate(seeds_pers):
                st.write(f"🔍 *{seed[:70]}*")
                try:
                    lotes = endo_bot._ddg(seed, max_s)
                    for l in lotes:
                        k = _slug(l.get("nome",""))
                        if not k or k in seen_sem: continue
                        seen_sem.add(k)
                        l["fonte_busca"] = seed
                        l["escopo"]      = pf or "ICP"
                        l["crm_status"]  = CRMAnalyzer.crm_status(l.get("nome",""), crm_df_all, l.get("cnpj",""))
                        l["lookalike"]   = CRMAnalyzer.lookalike_score(l)
                        # LinkedIn decisor
                        if enrich_li and not l.get("decisor"):
                            li_r = contact_hunter.linkedin_decisor(l.get("nome",""))
                            if li_r: l["decisor"] = li_r.get("decisor",""); time.sleep(0.5)
                        sem_results.append(l)
                except Exception as e:
                    st.write(f"⚠️ {e}")
                prog_s.progress((i+1)/len(seeds_pers))
                time.sleep(1.2)
            ss.update(label=f"✅ {len(sem_results)} leads ICP encontrados!", state="complete")
            st.session_state["persona_sem_results"] = sem_results

    sem_res = st.session_state.get("persona_sem_results", [])
    if sem_res:
        df_sem = pd.DataFrame(sem_res)
        CRM_ICON2 = {"vendida":"✅","perdida":"❌","em_andamento":"🔄","novo":"🆕","—":"🆕"}
        df_sem["CRM"] = df_sem.get("crm_status",pd.Series(["novo"]*len(df_sem))).fillna("novo").map(CRM_ICON2)
        df_sem["Lookalike"] = pd.to_numeric(df_sem.get("lookalike",pd.Series([0]*len(df_sem))),errors="coerce").fillna(0).astype(int).apply(lambda s: "🔥" if s>=70 else ("⭐" if s>=40 else ("➕" if s>0 else "—")))
        df_sem_sort = df_sem.sort_values("Lookalike", ascending=False, key=lambda x: x.map({"🔥":3,"⭐":2,"➕":1,"—":0}).fillna(0))
        cols_sem = ["nome","email","telefone","decisor","CRM","Lookalike","site","estado","cidade"]
        st.dataframe(df_sem_sort[[c for c in cols_sem if c in df_sem_sort.columns]].reset_index(drop=True),
                     use_container_width=True, height=360)
        st.markdown(f'<div style="font-size:.78rem;color:#888;margin:.3rem 0;">🔥 Alto = perfil similar às empresas ganhas no CRM &nbsp;·&nbsp; CRM ✅ = já é cliente</div>', unsafe_allow_html=True)
        if st.button("💾 Salvar todos na base Endotoxina", key="btn_save_sem", use_container_width=False):
            for lead in sem_res: db.upsert_endo(lead)
            st.success(f"✅ {len(sem_res)} leads salvos."); st.rerun()

    st.markdown("<hr style='border:1px solid #E0E0E0;margin:1rem 0;'>", unsafe_allow_html=True)

    # ════ SEÇÃO 5: ENRIQUECIMENTO INDIVIDUAL ═════════════════════════
    st.markdown('<div style="font-weight:700;font-size:.95rem;color:#2D2D2D;margin-bottom:.5rem;">⚡ Enriquecimento Multi-fonte de Lead Específico</div>', unsafe_allow_html=True)
    st.markdown('<div style="font-size:.82rem;color:#888;margin-bottom:.8rem;">Insira um CNPJ ou nome de empresa para buscar em todas as fontes disponíveis: Conodata, Casa dos Dados, LinkedIn, BrasilAPI.</div>', unsafe_allow_html=True)

    en1,en2,en3 = st.columns([2,2,1], gap="small")
    with en1: enrich_cnpj = st.text_input("CNPJ (somente números)", placeholder="00000000000000", key="enrich_cnpj_inp")
    with en2: enrich_nome = st.text_input("Nome da empresa (para LinkedIn)", placeholder="MAGNAMED TECNOLOGIA MEDICA", key="enrich_nome_inp")
    with en3:
        st.markdown("<div style='margin-top:1.6rem;'>", unsafe_allow_html=True)
        iniciar_enrich = st.button("⚡ Enriquecer", key="btn_enrich_single", use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

    if iniciar_enrich and (enrich_cnpj or enrich_nome):
        lead_enrich = {"cnpj": enrich_cnpj, "nome": enrich_nome}
        with st.status("⚡ Buscando em todas as fontes...", expanded=True) as se2:
            # BrasilAPI / ReceitaWS
            if enrich_cnpj:
                st.write("1️⃣ BrasilAPI / ReceitaWS...")
                lead_enrich = enricher.enrich_receita(lead_enrich)
                lead_enrich = enricher.enrich_contatos(lead_enrich)
            # Conodata
            if enrich_cnpj:
                st.write("2️⃣ Conodata...")
                cono = contact_hunter.conodata_cnpj(enrich_cnpj)
                for k,v in cono.items():
                    if v and not lead_enrich.get(k): lead_enrich[k]=v
            # LinkedIn decisor
            nome_busca = lead_enrich.get("nome","") or enrich_nome
            if nome_busca:
                st.write("3️⃣ LinkedIn (via DuckDuckGo)...")
                cargo_hint = " OR ".join(f'"{c}"' for c in ICP_PROFILE["cargos_decisores"][:5])
                li_r = contact_hunter.linkedin_decisor(nome_busca, cargo_hint)
                if li_r:
                    if not lead_enrich.get("decisor"): lead_enrich["decisor"] = li_r.get("decisor","")
                    st.write(f"  👤 Decisor encontrado: **{li_r.get('decisor','—')}** — {li_r.get('cargo','—')}")
            # CRM check
            lead_enrich["crm_status"] = CRMAnalyzer.crm_status(nome_busca, crm_df_all, re.sub(r"\D","",enrich_cnpj or ""))
            lead_enrich["lookalike"]  = CRMAnalyzer.lookalike_score(lead_enrich)
            se2.update(label="✅ Enriquecimento concluído!", state="complete")

        # Exibir resultado estruturado
        st.markdown("<br>", unsafe_allow_html=True)
        campos = [
            ("🏢 Empresa",    lead_enrich.get("nome","—")),
            ("🪪 CNPJ",       lead_enrich.get("cnpj","—")),
            ("✉️ E-mail",      lead_enrich.get("email","—")),
            ("📞 Telefone",   lead_enrich.get("telefone","—")),
            ("👤 Decisor",    lead_enrich.get("decisor","—")),
            ("🏭 Setor/CNAE", lead_enrich.get("setor","—")),
            ("📍 Cidade/UF",  f'{lead_enrich.get("cidade","—")} / {lead_enrich.get("estado","—")}'),
            ("🔗 CRM Status", lead_enrich.get("crm_status","—")),
            ("⭐ Lookalike",  f'{lead_enrich.get("lookalike",0)}/100'),
            ("📋 Situação",   lead_enrich.get("situacao","—")),
        ]
        col_ea, col_eb = st.columns(2, gap="medium")
        for i,(lbl,val) in enumerate(campos):
            with (col_ea if i%2==0 else col_eb):
                is_found = val and val not in ("—","None","")
                bg = "#E0F4F4" if is_found else "#F9F9F9"
                border = "#0D9291" if is_found else "#E0E0E0"
                st.markdown(f'<div style="background:{bg};border:1px solid {border};border-radius:8px;padding:.5rem .8rem;margin-bottom:.4rem;"><span style="font-size:.75rem;font-weight:600;color:#888;">{lbl}</span><div style="font-size:.9rem;color:#2D2D2D;font-weight:500;margin-top:1px;">{val}</div></div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════
#  ABA 6 — DASHBOARD INTERATIVO
# ══════════════════════════════════════════════════════════════════════
with t7:
    _dash_mod = "Laboratório" if IS_LAB else "Certificadora OCP"
    st.markdown(
        f'<div class="sec-title">📊 Painel — {_dash_mod}</div>',
        unsafe_allow_html=True
    )
    st.markdown(
        f'<div class="sec-desc">Métricas de conversão · Distribuição geográfica · Performance CRM · Clique nos cards para filtrar.</div>',
        unsafe_allow_html=True
    )

    port_stat = pid_cod if IS_OCP else None
    try:    stats = db.get_stats(port_stat)
    except: stats = {"total":0,"hot":0,"vencido":0,"contatado":0,"medio":0,"normal":0}
    endo_st2  = db.get_endo_stats()
    recon_n   = len(db.get_recontatar(port_stat))

    # ── Cards clicáveis via botões Streamlit ──────────────────────────
    st.markdown('<div style="font-size:.78rem;color:#888;margin-bottom:.5rem;">💡 Clique em um card para filtrar os leads correspondentes abaixo.</div>', unsafe_allow_html=True)

    kc1,kc2,kc3,kc4,kc5,kc6 = st.columns(6,gap="small")
    filter_map = {
        "todos":     {"portaria":port_stat},
        "hot":       {"portaria":port_stat,"urgencia":"hot"},
        "vencido":   {"portaria":port_stat,"urgencia":"vencido"},
        "contatado": {"portaria":port_stat,"status_crm":"contatado"},
        "medio":     {"portaria":port_stat,"urgencia":"medio"},
        "normal":    {"portaria":port_stat,"urgencia":"normal"},
    }
    active = st.session_state.get("dash_filter")

    with kc1:
        st.markdown(f'<div class="mc{"  active" if active=="todos" else ""}"><div class="ml">INMETRO</div><div class="mn">{stats["total"]}</div></div>', unsafe_allow_html=True)
        if st.button("Ver todos", key="d_todos", use_container_width=True):
            st.session_state["dash_filter"]="todos"; st.session_state["dash_leads"]=db.get_leads(filter_map["todos"] or {}); st.rerun()
    with kc2:
        st.markdown(f'<div class="mc{"  active" if active=="hot" else ""}"><div class="ml">Quentes 🔥</div><div class="mn mn-hot">{stats["hot"]}</div></div>', unsafe_allow_html=True)
        if st.button("Ver quentes", key="d_hot", use_container_width=True):
            st.session_state["dash_filter"]="hot"; st.session_state["dash_leads"]=db.get_leads(filter_map["hot"] or {}); st.rerun()
    with kc3:
        st.markdown(f'<div class="mc{"  active" if active=="vencido" else ""}"><div class="ml">Vencidos ⚠️</div><div class="mn" style="color:#6B7280;">{stats["vencido"]}</div></div>', unsafe_allow_html=True)
        if st.button("Ver vencidos", key="d_venc", use_container_width=True):
            st.session_state["dash_filter"]="vencido"; st.session_state["dash_leads"]=db.get_leads(filter_map["vencido"] or {}); st.rerun()
    with kc4:
        st.markdown(f'<div class="mc{"  active" if active=="contatado" else ""}"><div class="ml">Contatados ✔️</div><div class="mn mn-am">{stats["contatado"]}</div></div>', unsafe_allow_html=True)
        if st.button("Ver contatados", key="d_cont", use_container_width=True):
            st.session_state["dash_filter"]="contatado"; st.session_state["dash_leads"]=db.get_leads(filter_map["contatado"] or {}); st.rerun()
    with kc5:
        st.markdown(f'<div class="mc{"  active" if active=="endotoxina" else ""}"><div class="ml">Endotoxina 🧬</div><div class="mn mn-vl">{endo_st2["total"]}</div></div>', unsafe_allow_html=True)
        if st.button("Ver endotoxina", key="d_endo", use_container_width=True):
            st.session_state["dash_filter"]="endotoxina"; st.session_state["dash_leads"]=None; st.rerun()
    with kc6:
        st.markdown(f'<div class="mc{"  active" if active=="medio" else ""}"><div class="ml">Médio Prazo ⏳</div><div class="mn mn-am">{stats["medio"]}</div></div>', unsafe_allow_html=True)
        if st.button("Ver médio prazo", key="d_med", use_container_width=True):
            st.session_state["dash_filter"]="medio"; st.session_state["dash_leads"]=db.get_leads(filter_map["medio"] or {}); st.rerun()

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Análise CRM inline no dashboard ──────────────────────────────
    if crm_df_all is not None and "Estado" in crm_df_all.columns:
        crm_vendidas = (crm_df_all["Estado"]=="Vendida").sum()
        crm_perdidas = (crm_df_all["Estado"]=="Perdida").sum()
        crm_andamento= (crm_df_all["Estado"]=="Em Andamento").sum()
        crm_conv = round(crm_vendidas/(crm_vendidas+crm_perdidas)*100) if (crm_vendidas+crm_perdidas)>0 else 0
        st.markdown('<div style="font-weight:600;font-size:.88rem;color:#2D2D2D;margin-bottom:.5rem;">📊 Performance CRM</div>', unsafe_allow_html=True)
        dc1,dc2,dc3,dc4 = st.columns(4,gap="small")
        with dc1: st.markdown(f'<div class="mc"><div class="ml">Total CRM</div><div class="mn mn-bl">{len(crm_df_all)}</div></div>', unsafe_allow_html=True)
        with dc2: st.markdown(f'<div class="mc"><div class="ml">Ganhos ✅</div><div class="mn" style="color:#16a34a;">{crm_vendidas}</div></div>', unsafe_allow_html=True)
        with dc3: st.markdown(f'<div class="mc"><div class="ml">Perdidos ❌</div><div class="mn mn-hot">{crm_perdidas}</div></div>', unsafe_allow_html=True)
        with dc4: st.markdown(f'<div class="mc"><div class="ml">Conversão</div><div class="mn" style="color:#0D9291;">{crm_conv}%</div></div>', unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)

    # ── Detalhamento por distribuição ─────────────────────────────────
    col_g1,col_g2 = st.columns(2, gap="medium")
    with col_g1:
        st.markdown('<div style="font-weight:600;font-size:.88rem;color:#2D2D2D;margin-bottom:.5rem;">📍 Leads por Estado</div>', unsafe_allow_html=True)
        dist_est = db.get_dist_estado(port_stat)
        if dist_est:
            total_est = sum(dist_est.values())
            for uf,n in list(dist_est.items())[:8]:
                pct = round(n/total_est*100)
                st.markdown(f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:5px;"><span style="font-size:.8rem;font-weight:600;min-width:28px;">{uf}</span><div class="progress-bar-wrap" style="flex:1;"><div class="progress-bar" style="width:{pct}%;background:#0D9291;"></div></div><span style="font-size:.78rem;color:#888;min-width:40px;text-align:right;">{n} ({pct}%)</span></div>', unsafe_allow_html=True)
        else:
            st.caption("Nenhum dado disponível")

    with col_g2:
        st.markdown('<div style="font-weight:600;font-size:.88rem;color:#2D2D2D;margin-bottom:.5rem;">🏢 Top OCPs</div>', unsafe_allow_html=True)
        dist_ocp = db.get_dist_ocp(port_stat)
        if dist_ocp:
            total_ocp = sum(dist_ocp.values())
            for ocp,n in list(dist_ocp.items())[:8]:
                pct = round(n/total_ocp*100)
                st.markdown(f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:5px;"><span style="font-size:.78rem;min-width:130px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;" title="{ocp}">{ocp[:18]}</span><div class="progress-bar-wrap" style="flex:1;"><div class="progress-bar" style="width:{pct}%;background:#3B82F6;"></div></div><span style="font-size:.78rem;color:#888;min-width:40px;text-align:right;">{n}</span></div>', unsafe_allow_html=True)
        else:
            st.caption("Nenhum dado disponível")

    # ── Tabela filtrada pelo card clicado ─────────────────────────────
    dash_leads_show = st.session_state.get("dash_leads")
    dash_filter_act = st.session_state.get("dash_filter")

    if dash_filter_act == "endotoxina":
        st.markdown("<hr style='border:1px solid #E0E0E0;margin:1rem 0;'>", unsafe_allow_html=True)
        st.markdown(f'<div style="font-weight:600;font-size:.88rem;margin-bottom:.5rem;">🧬 Leads Endotoxina ({endo_st2["total"]})</div>', unsafe_allow_html=True)
        endo_all = db.get_endo_leads()
        if endo_all:
            df_e=pd.DataFrame(endo_all)
            st.dataframe(df_e[["nome","email","telefone","site","estado","cidade","status"]], use_container_width=True, height=300)
        else:
            st.caption("Nenhum lead de endotoxina.")

    elif dash_leads_show is not None:
        label_map = {"todos":"Todos os Leads","hot":"🔥 Leads Quentes","vencido":"⚠️ Vencidos",
                     "contatado":"✔️ Contatados","medio":"⏳ Médio Prazo"}
        label = label_map.get(dash_filter_act, dash_filter_act)
        st.markdown("<hr style='border:1px solid #E0E0E0;margin:1rem 0;'>", unsafe_allow_html=True)
        st.markdown(f'<div style="font-weight:600;font-size:.88rem;margin-bottom:.5rem;">{label} ({len(dash_leads_show)})</div>', unsafe_allow_html=True)
        if dash_leads_show:
            df_d=pd.DataFrame(dash_leads_show)
            urgencia_color = {"hot":"🔥","vencido":"⚠️","medio":"⏳","normal":"✅"}
            df_d["🏷"]=df_d["urgencia"].map(urgencia_color).fillna("—")
            cols_d=["🏷","nome","cnpj","email","telefone","decisor","ocp","validade","estado","cidade","status_crm"]
            st.dataframe(df_d[[c for c in cols_d if c in df_d.columns]], use_container_width=True, height=320)
        else:
            st.caption("Nenhum lead nesta categoria.")

    # ── Exportação ────────────────────────────────────────────────────
    st.markdown("<hr style='border:1px solid #E0E0E0;margin:1rem 0;'>", unsafe_allow_html=True)
    st.markdown('<div style="font-weight:600;font-size:.9rem;color:#2D2D2D;margin-bottom:.5rem;">📥 Exportar Relatórios</div>', unsafe_allow_html=True)
    col_xa,col_xb,_ = st.columns([1,1,2],gap="medium")
    with col_xa:
        if IS_LAB:
            st.info("Selecione portaria INMETRO para exportar.")
        else:
            if st.button("📊 Gerar Excel INMETRO", key="btn_exp_inm", use_container_width=True):
                with st.spinner("Gerando..."):
                    file_inm=exporter.generate_excel(portaria_id)
                st.download_button("⬇  Baixar INMETRO", data=file_inm,
                                   file_name=f"prospec_scitec_{portaria_id}.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with col_xb:
        if st.button("📊 Gerar Excel Endotoxina", key="btn_exp_endo", use_container_width=True):
            with st.spinner("Gerando..."):
                file_endo=exporter.generate_endo_excel()
            st.download_button("⬇  Baixar Endotoxina", data=file_endo,
                               file_name="leads_endotoxina.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")